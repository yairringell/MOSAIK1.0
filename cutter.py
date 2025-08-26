import sys
import csv
import os
import numpy as np
try:
    import cv2
except ImportError:
    print("Warning: OpenCV not installed. Blob detection will use fallback method.")
    cv2 = None
try:
    import ezdxf
except ImportError:
    print("Warning: ezdxf not installed. DXF export will be disabled.")
    ezdxf = None
import xml.etree.ElementTree as ET
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QFileDialog, QGraphicsView, 
                             QGraphicsScene, QGraphicsPixmapItem, QMenuBar, QAction,
                             QGraphicsRectItem, QGraphicsPolygonItem, QGraphicsTextItem,
                             QGraphicsLineItem, QGraphicsEllipseItem)
from PyQt5.QtCore import Qt, QRectF, QPointF, QSize, QTimer
from PyQt5.QtGui import QColor, QPen, QBrush, QPixmap, QPolygonF, QPainterPath, QPainter, QFont, QFontMetrics

class ScaleBar(QWidget):
    """Custom scale bar widget that shows pixel measurements"""
    def __init__(self, orientation='horizontal', parent=None):
        super().__init__(parent)
        self.orientation = orientation  # 'horizontal' or 'vertical'
        self.scale_factor = 1.0  # Current zoom scale factor
        self.scene_rect = QRectF(0, 0, 1000, 1000)  # Scene dimensions
        self.view_rect = QRectF(0, 0, 500, 500)  # Visible area
        
        # Setup appearance
        self.setAutoFillBackground(True)
        palette = self.palette()
        palette.setColor(palette.Window, QColor(240, 240, 240))
        self.setPalette(palette)
        
        # Set size based on orientation
        if orientation == 'horizontal':
            self.setFixedHeight(30)
            self.setMinimumWidth(100)
        else:
            self.setFixedWidth(30)
            self.setMinimumHeight(100)
    
    def update_scale(self, scale_factor, scene_rect, view_rect):
        """Update the scale bar based on current zoom and view"""
        self.scale_factor = scale_factor
        self.scene_rect = scene_rect
        self.view_rect = view_rect
        self.update()  # Trigger repaint
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Set font
        font = QFont()
        font.setPointSize(8)
        painter.setFont(font)
        
        # Calculate tick spacing
        if self.orientation == 'horizontal':
            self._paint_horizontal_scale(painter)
        else:
            self._paint_vertical_scale(painter)
    
    def _paint_horizontal_scale(self, painter):
        """Paint horizontal scale bar"""
        width = self.width()
        height = self.height()
        
        # Calculate pixel size in scene coordinates
        pixels_per_scene_unit = 1.0 / self.scale_factor
        
        # Determine appropriate tick spacing (in pixels)
        tick_spacings = [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000]
        target_tick_width = 50  # Desired width between major ticks in pixels
        
        # Find best tick spacing
        best_spacing = tick_spacings[0]
        for spacing in tick_spacings:
            scene_spacing = spacing * pixels_per_scene_unit
            pixel_width = scene_spacing * self.scale_factor
            if pixel_width >= target_tick_width:
                best_spacing = spacing
                break
            best_spacing = spacing
        
        # Draw background
        painter.fillRect(0, 0, width, height, QColor(240, 240, 240))
        
        # Draw border
        painter.setPen(QPen(QColor(100, 100, 100), 1))
        painter.drawLine(0, height-1, width, height-1)
        
        # Calculate visible range in scene coordinates
        view_left = self.view_rect.left()
        view_right = self.view_rect.right()
        
        # Calculate starting tick position
        start_tick = int(view_left // best_spacing) * best_spacing
        
        # Draw ticks and labels
        painter.setPen(QPen(QColor(50, 50, 50), 1))
        
        current_tick = start_tick
        while current_tick <= view_right + best_spacing:
            # Convert scene position to widget position
            scene_pos = current_tick - view_left
            widget_pos = (scene_pos / (view_right - view_left)) * width
            
            if 0 <= widget_pos <= width:
                # Draw major tick
                painter.drawLine(int(widget_pos), height-10, int(widget_pos), height-1)
                
                # Draw label
                label = str(int(current_tick))
                fm = QFontMetrics(painter.font())
                label_width = fm.width(label)
                label_x = int(widget_pos - label_width/2)
                painter.drawText(label_x, 12, label)
                
                # Draw minor ticks
                minor_spacing = best_spacing / 5
                for i in range(1, 5):
                    minor_pos = current_tick + i * minor_spacing
                    if minor_pos <= view_right:
                        minor_scene_pos = minor_pos - view_left
                        minor_widget_pos = (minor_scene_pos / (view_right - view_left)) * width
                        if 0 <= minor_widget_pos <= width:
                            painter.drawLine(int(minor_widget_pos), height-5, int(minor_widget_pos), height-1)
            
            current_tick += best_spacing
    
    def _paint_vertical_scale(self, painter):
        """Paint vertical scale bar"""
        width = self.width()
        height = self.height()
        
        # Calculate pixel size in scene coordinates
        pixels_per_scene_unit = 1.0 / self.scale_factor
        
        # Determine appropriate tick spacing (in pixels)
        tick_spacings = [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000]
        target_tick_height = 50  # Desired height between major ticks in pixels
        
        # Find best tick spacing
        best_spacing = tick_spacings[0]
        for spacing in tick_spacings:
            scene_spacing = spacing * pixels_per_scene_unit
            pixel_height = scene_spacing * self.scale_factor
            if pixel_height >= target_tick_height:
                best_spacing = spacing
                break
            best_spacing = spacing
        
        # Draw background
        painter.fillRect(0, 0, width, height, QColor(240, 240, 240))
        
        # Draw border
        painter.setPen(QPen(QColor(100, 100, 100), 1))
        painter.drawLine(width-1, 0, width-1, height)
        
        # Calculate visible range in scene coordinates
        view_top = self.view_rect.top()
        view_bottom = self.view_rect.bottom()
        
        # Calculate starting tick position
        start_tick = int(view_top // best_spacing) * best_spacing
        
        # Draw ticks and labels
        painter.setPen(QPen(QColor(50, 50, 50), 1))
        
        current_tick = start_tick
        while current_tick <= view_bottom + best_spacing:
            # Convert scene position to widget position
            scene_pos = current_tick - view_top
            widget_pos = (scene_pos / (view_bottom - view_top)) * height
            
            if 0 <= widget_pos <= height:
                # Draw major tick
                painter.drawLine(width-10, int(widget_pos), width-1, int(widget_pos))
                
                # Draw label (rotated for vertical)
                painter.save()
                painter.translate(8, int(widget_pos))
                painter.rotate(-90)
                label = str(int(current_tick))
                fm = QFontMetrics(painter.font())
                label_width = fm.width(label)
                painter.drawText(-label_width//2, 0, label)
                painter.restore()
                
                # Draw minor ticks
                minor_spacing = best_spacing / 5
                for i in range(1, 5):
                    minor_pos = current_tick + i * minor_spacing
                    if minor_pos <= view_bottom:
                        minor_scene_pos = minor_pos - view_top
                        minor_widget_pos = (minor_scene_pos / (view_bottom - view_top)) * height
                        if 0 <= minor_widget_pos <= height:
                            painter.drawLine(width-5, int(minor_widget_pos), width-1, int(minor_widget_pos))
            
            current_tick += best_spacing

class GridHandle(QGraphicsRectItem):
    """Draggable handle for moving the grid"""
    def __init__(self, parent_view):
        super().__init__(0, 0, 20, 20)  # 20x20 pixel handle
        self.parent_view = parent_view
        
        # Set appearance - red rectangle
        self.setPen(QPen(QColor(255, 0, 0), 1))  # Red border
        self.setBrush(QBrush(QColor(255, 0, 0, 150)))  # Semi-transparent red fill
        
        # Make it movable
        self.setFlag(QGraphicsRectItem.ItemIsMovable, True)
        self.setFlag(QGraphicsRectItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsRectItem.ItemSendsGeometryChanges, True)
        
        # Set Z-value to be in front of grid but behind shapes
        self.setZValue(-0.3)
    
    def itemChange(self, change, value):
        """Handle position changes to move the entire grid"""
        if change == QGraphicsRectItem.ItemPositionChange and self.scene():
            # Calculate the new grid offset
            new_pos = value
            self.parent_view.grid_offset_x = new_pos.x()
            self.parent_view.grid_offset_y = new_pos.y()
            
            # Update grid position
            self.parent_view.update_grid_position()
        
        return super().itemChange(change, value)

class ScalableRectangle(QGraphicsRectItem):
    """Simplified rectangle class for display only"""
    def __init__(self, x, y, width, height, initial_color=None):
        super().__init__(0, 0, width, height)  # Create rect at origin
        self.setPos(x, y)  # Set position
        
        # Set flags to make non-interactive for viewing
        self.setFlag(QGraphicsRectItem.ItemIsMovable, False)
        self.setFlag(QGraphicsRectItem.ItemIsSelectable, False)
        
        # Set appearance - black frame with minimal thickness
        self.setPen(QPen(QColor(0, 0, 0), 0))  # Black frame, minimal thickness
        self.setBrush(QBrush(Qt.transparent))  # Start transparent
        
        # Store properties
        self.current_rotation = 0
        self.is_filled = False
        self.fill_color = Qt.transparent
        self.serial_number = 0
        
        # Set rotation center to center of rectangle
        rect_center = self.rect().center()
        self.setTransformOriginPoint(rect_center)
    
    def set_fill_color(self, color):
        """Set fill color"""
        if color and color.isValid():
            self.fill_color = color
            self.is_filled = True
            self.setBrush(QBrush(color))
        else:
            self.is_filled = False
            self.setBrush(QBrush(Qt.transparent))

class ScalableTriangle(QGraphicsPolygonItem):
    """Simplified triangle class for display only"""
    def __init__(self, x, y, size, initial_color=None):
        # Create a 90-degree right triangle
        triangle_points = [
            QPointF(0, 0),
            QPointF(size, 0),
            QPointF(0, size)
        ]
        triangle_polygon = QPolygonF(triangle_points)
        
        super().__init__(triangle_polygon)
        self.setPos(x, y)
        
        # Set flags to make non-interactive for viewing
        self.setFlag(QGraphicsPolygonItem.ItemIsMovable, False)
        self.setFlag(QGraphicsPolygonItem.ItemIsSelectable, False)
        
        # Set appearance - black frame with minimal thickness
        self.setPen(QPen(QColor(0, 0, 0), 0))  # Black frame, minimal thickness
        self.setBrush(QBrush(Qt.transparent))  # Start transparent
        
        # Store properties
        self.current_rotation = 0
        self.is_filled = False
        self.fill_color = Qt.transparent
        self.serial_number = 0
        self.size = size
        
        # Set rotation center to center of triangle height (geometric center)
        # For a right triangle with vertices at (0,0), (size,0), (0,size)
        # The centroid is at (size/3, size/3)
        triangle_center = QPointF(size/3, size/3)
        self.setTransformOriginPoint(triangle_center)
    
    def set_fill_color(self, color):
        """Set fill color"""
        if color and color.isValid():
            self.fill_color = color
            self.is_filled = True
            self.setBrush(QBrush(color))
        else:
            self.is_filled = False
            self.setBrush(QBrush(Qt.transparent))

class CutterView(QGraphicsView):
    """Graphics view with zoom capabilities"""
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Create scene
        self.scene = QGraphicsScene()
        self.setScene(self.scene)
        
        # Set up zooming parameters - copied from tessera1_2.py
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)
        
        # Enable drag mode for panning
        self.setDragMode(QGraphicsView.RubberBandDrag)
        
        # Set scene rect to match tessera1_2.py positioning
        self.scene.setSceneRect(QRectF(0, 0, 2000, 1100))
        
        # Enable mouse tracking
        self.setMouseTracking(True)
        
        # Background
        self.background_item = None
        
        # Grid items for 6x6 grid of 250x250 boxes
        self.grid_items = []
        self.grid_labels = []  # Store grid labels separately
        self.cut_lines = []  # Store cut lines
        self.grid_visible = False
        self.grid_handle = None
        self.grid_offset_x = 0
        self.grid_offset_y = 0
        
        # Store inclusion data for consistent numbering and debugging
        self.box_inclusion_data = {}  # Dictionary to store which shapes belong to which boxes
        
        # Shape number display tracking
        self.numbers_visible = False
        self.number_text_items = []  # List to store text items showing shape numbers
        self.placed_ref_shapes = []  # List to store bounds of placed reference shapes
        
        # Create scale bars
        self.horizontal_scale_bar = ScaleBar('horizontal', self)
        self.vertical_scale_bar = ScaleBar('vertical', self)
        
        # Position scale bars (they will be repositioned in resizeEvent)
        self.horizontal_scale_bar.move(30, 0)
        self.vertical_scale_bar.move(0, 30)
        
        # Set viewport margins to account for scale bars
        self.setViewportMargins(30, 30, 0, 0)
        
        # Connect to view change events to update scale bars
        self.horizontalScrollBar().valueChanged.connect(self.update_scale_bars)
        self.verticalScrollBar().valueChanged.connect(self.update_scale_bars)
        
        # Also connect to scene rect changes
        self.scene.sceneRectChanged.connect(self.update_scale_bars)
        
    def wheelEvent(self, event):
        """Handle mouse wheel for zooming towards cursor position"""
        # Zoom factor
        zoomInFactor = 1.15
        zoomOutFactor = 1 / zoomInFactor
        
        # Save the scene pos
        oldPos = self.mapToScene(event.pos())
        
        # Zoom
        if event.angleDelta().y() > 0:
            zoomFactor = zoomInFactor
        else:
            zoomFactor = zoomOutFactor
        self.scale(zoomFactor, zoomFactor)
        
        # Get the new position and move scene to old position
        newPos = self.mapToScene(event.pos())
        delta = newPos - oldPos
        self.translate(delta.x(), delta.y())
        
        # Update scale bars after zooming
        QTimer.singleShot(50, self.update_scale_bars)
    
    def set_background_image(self, pixmap):
        """Set background image"""
        if self.background_item:
            self.scene.removeItem(self.background_item)
        
        self.background_item = QGraphicsPixmapItem(pixmap)
        self.background_item.setZValue(-1)  # Put background behind everything
        self.scene.addItem(self.background_item)
    
    def add_shape(self, shape):
        """Add a shape to the scene"""
        self.scene.addItem(shape)
    
    def clear_shapes(self):
        """Clear all shapes but preserve background items"""
        # First hide shape numbers if they are visible
        if self.numbers_visible:
            self.hide_shape_numbers()
        
        # Remove all items except background, grid, grid labels, cut lines, and grid handle
        items_to_remove = []
        for item in self.scene.items():
            if (item != self.background_item and 
                item not in self.grid_items and 
                item not in self.grid_labels and
                item not in self.cut_lines and
                item != self.grid_handle):
                items_to_remove.append(item)
        
        # Remove items
        for item in items_to_remove:
            self.scene.removeItem(item)
        
        print(f"Cleared {len(items_to_remove)} shape items")
    
    def toggle_shape_numbers(self):
        """Toggle display of shape serial numbers on shapes"""
        if self.numbers_visible:
            self.hide_shape_numbers()
        else:
            self.show_shape_numbers()
    
    def show_shape_numbers(self):
        """Display array position numbers on all shapes and draw first 10 shapes in order"""
        if self.numbers_visible:
            return  # Already showing numbers
        
        self.numbers_visible = True
        
        # Get all shapes in the scene and sort them by Y position (same as CSV array order)
        shapes = []
        for item in self.scene.items():
            if hasattr(item, 'serial_number') and isinstance(item, (ScalableRectangle, ScalableTriangle)):
                shapes.append(item)
        
        # Sort shapes by their Y position (top to bottom) - same order as CSV array
        shapes.sort(key=lambda item: item.pos().y())
        
        # Display array position number (1-based) on each shape
        for index, shape in enumerate(shapes):
            array_position = index + 1  # 1-based numbering
            
            # Restore original colors for the shape
            original_fill_color = getattr(shape, 'original_fill_color', '')
            original_frame_color = getattr(shape, 'original_frame_color', '#8B4513')
            original_is_filled = getattr(shape, 'original_is_filled', False)
            
            # Set frame color
            frame_color = QColor(original_frame_color) if original_frame_color else QColor(139, 69, 19)
            shape.setPen(QPen(frame_color, 0))
            
            # Set fill color
            if original_is_filled and original_fill_color:
                fill_color = QColor(original_fill_color)
                shape.setBrush(QBrush(fill_color))
            else:
                shape.setBrush(QBrush(Qt.transparent))
            
            # Create text item for this shape's array position
            text_item = QGraphicsTextItem(str(array_position))
            
            # Set very small font
            font = QFont()
            font.setPointSize(1)  # Smallest possible font
            font.setBold(True)
            text_item.setFont(font)
            
            # Set text color to black for visibility
            text_item.setDefaultTextColor(QColor(0, 0, 0))
            
            # Position text at the center of the shape
            shape_rect = shape.sceneBoundingRect()
            text_rect = text_item.boundingRect()
            
            # Calculate center position based on shape type
            if isinstance(shape, ScalableTriangle):
                # For triangles, use geometric center (centroid)
                triangle_size = getattr(shape, 'size', 10)
                # Get shape position and add centroid offset
                shape_pos = shape.pos()
                center_x = shape_pos.x() + triangle_size/3
                center_y = shape_pos.y() + triangle_size/3
            else:
                # For rectangles, use bounding rect center
                center_x = shape_rect.center().x()
                center_y = shape_rect.center().y()
            
            # Center the text on the calculated center
            text_x = center_x - text_rect.width() / 2
            text_y = center_y - text_rect.height() / 2
            text_item.setPos(text_x, text_y)
            
            # Set high z-value to ensure text appears on top
            text_item.setZValue(100)
            
            # Add to scene and track it
            self.scene.addItem(text_item)
            self.number_text_items.append(text_item)
        
        # Draw first 10 shapes in array order starting at (10,10) with 25px spacing
        self.draw_array_reference(shapes)
        
        print(f"Showing array position numbers on {len(self.number_text_items)} shapes")
    
    def draw_array_reference(self, first_shapes):
        """Draw reference shapes: first half at (10,10), second half at (10,400), both color-grouped"""
        start_center_x = 25
        start_center_y = 40
        initial_spacing = 10  # Start with 10 pixels spacing
        max_x = 595  # Maximum x position before starting new row
        row_spacing = 23  # Distance between rows
        
        # Add frame rectangle from (0,0) to (620,515)
        frame_rect = QGraphicsRectItem(0, 0, 620, 515)
        frame_rect.setPen(QPen(QColor(0, 0, 0), 2))  # Black border, 2px thick
        frame_rect.setBrush(QBrush())  # No fill, transparent
        frame_rect.setZValue(-10)  # Behind other items
        self.scene.addItem(frame_rect)
        self.number_text_items.append(frame_rect)  # Track for removal
        
        # Keep track of placed shapes for collision detection
        self.placed_ref_shapes = []
        
        # Function to group all shapes by color
        def group_shapes_by_color(shapes):
            color_groups = {}
            for i, shape in enumerate(shapes):
                # Get the original fill color for grouping
                fill_color = getattr(shape, 'original_fill_color', '')
                frame_color = getattr(shape, 'original_frame_color', '#8B4513')
                is_filled = getattr(shape, 'original_is_filled', False)
                
                # Use fill color if filled, otherwise use frame color for grouping
                group_color = fill_color if (is_filled and fill_color) else frame_color
                
                if group_color not in color_groups:
                    color_groups[group_color] = []
                color_groups[group_color].append((i, shape))
            
            # Sort color groups for consistent ordering and flatten
            sorted_color_groups = sorted(color_groups.items())
            result = []
            for group_color, shapes_in_group in sorted_color_groups:
                for original_index, shape in shapes_in_group:
                    result.append((original_index, shape))
            return result
        
        # Group all shapes by color and display starting at (10,10)
        all_shapes_grouped = group_shapes_by_color(first_shapes)
        current_x = start_center_x
        current_y = start_center_y
        
        for original_index, original_shape in all_shapes_grouped:
            # Find a safe position that doesn't overlap with previous shapes
            safe_x = self._find_safe_x_position(original_shape, current_x, current_y, max_x)
            
            self._draw_single_reference_shape(original_shape, original_index, safe_x, current_y)
            
            # Update position for next shape
            current_x = safe_x + self._get_shape_width(original_shape) + 5  # Add 5px buffer
            
            # Check if we need to start a new row
            if current_x > max_x:
                current_x = start_center_x
                current_y += row_spacing
                self.placed_ref_shapes.clear()  # Clear for new row
    
    def _find_safe_x_position(self, shape, start_x, y, max_x):
        """Find a safe X position that doesn't overlap with existing shapes or original canvas shapes"""
        test_x = start_x
        shape_bounds = self._get_shape_bounds(shape, test_x, y)
        
        # Check against all shapes placed on this row AND original canvas shapes
        max_attempts = 100
        attempts = 0
        
        while attempts < max_attempts:
            collision = False
            shape_bounds = self._get_shape_bounds(shape, test_x, y)
            
            # Check collision with other reference shapes on this row
            for existing_bounds in self.placed_ref_shapes:
                if self._bounds_overlap(shape_bounds, existing_bounds):
                    collision = True
                    break
            
            # Also check collision with reserved rectangle area (160,120) to (440,390)
            if not collision:
                reserved_left = 160
                reserved_top = 120
                reserved_right = 440
                reserved_bottom = 390
                
                # Check if reference shape would overlap with reserved area
                if (shape_bounds['right'] > reserved_left and 
                    shape_bounds['left'] < reserved_right and
                    shape_bounds['bottom'] > reserved_top and 
                    shape_bounds['top'] < reserved_bottom):
                    collision = True
            
            if not collision:
                # Store this shape's bounds for future collision checks
                self.placed_ref_shapes.append(shape_bounds)
                return test_x
            
            # Move 3 pixels to the right and try again
            test_x += 3
            attempts += 1
            
            # If we've gone too far, break (let main logic handle row wrapping)
            if test_x > max_x:
                break
        
        # If we couldn't find a safe position, return the start position
        # and let the main logic handle row wrapping
        return start_x
    
    def _get_shape_width(self, shape):
        """Get the effective width of a shape including rotation"""
        if isinstance(shape, ScalableRectangle):
            orig_rect = shape.rect()
            width = orig_rect.width()
            height = orig_rect.height()
            rotation = getattr(shape, 'current_rotation', 0)
            
            if abs(rotation) > 5:  # If significantly rotated
                # Use diagonal as safe width
                return (width**2 + height**2)**0.5
            else:
                return width
        else:  # ScalableTriangle
            orig_size = getattr(shape, 'size', 10)
            rotation = getattr(shape, 'current_rotation', 0)
            
            if abs(rotation) > 5:  # If significantly rotated
                return orig_size * 1.414  # sqrt(2)
            else:
                return orig_size
    
    def _get_shape_bounds(self, shape, center_x, center_y):
        """Calculate the bounding box of a shape at a given position"""
        if isinstance(shape, ScalableRectangle):
            orig_rect = shape.rect()
            width = orig_rect.width()
            height = orig_rect.height()
            rotation = getattr(shape, 'current_rotation', 0)
            
            if abs(rotation) > 5:  # If significantly rotated
                # Use diagonal for safe bounds
                diagonal = (width**2 + height**2)**0.5
                half_diagonal = diagonal / 2
                return {
                    'left': center_x - half_diagonal,
                    'right': center_x + half_diagonal,
                    'top': center_y - half_diagonal,
                    'bottom': center_y + half_diagonal
                }
            else:
                return {
                    'left': center_x - width/2,
                    'right': center_x + width/2,
                    'top': center_y - height/2,
                    'bottom': center_y + height/2
                }
        else:  # ScalableTriangle
            orig_size = getattr(shape, 'size', 10)
            rotation = getattr(shape, 'current_rotation', 0)
            
            if abs(rotation) > 5:  # If significantly rotated
                # Use expanded bounds for rotated triangle
                expanded_size = orig_size * 1.414  # sqrt(2)
                half_size = expanded_size / 2
                return {
                    'left': center_x - half_size,
                    'right': center_x + half_size,
                    'top': center_y - half_size,
                    'bottom': center_y + half_size
                }
            else:
                half_size = orig_size / 2
                return {
                    'left': center_x - half_size,
                    'right': center_x + half_size,
                    'top': center_y - half_size,
                    'bottom': center_y + half_size
                }
    
    def _bounds_overlap(self, bounds1, bounds2):
        """Check if two bounding boxes overlap"""
        return not (bounds1['right'] <= bounds2['left'] or 
                   bounds1['left'] >= bounds2['right'] or
                   bounds1['bottom'] <= bounds2['top'] or
                   bounds1['top'] >= bounds2['bottom'])
    
    def _find_non_overlapping_position(self, shape, start_x, start_y, initial_spacing, max_x):
        """Find a position where the shape won't overlap with existing shapes"""
        test_x = start_x
        test_y = start_y
        
        # Calculate the bounding box of the shape at the test position
        bounds = self._get_shape_bounds(shape, test_x, test_y)
        
        # Check for overlaps with existing shapes, but only move horizontally
        # Don't let collision detection force a new row - that should be handled by the main layout logic
        max_attempts = 50  # Prevent infinite loops
        attempts = 0
        
        while self._check_overlap_with_existing(bounds) and attempts < max_attempts:
            # Move 2 pixels to the right and try again
            test_x += 2
            attempts += 1
            
            # If we've moved too far right, break and let the main logic handle row wrapping
            if test_x > max_x:
                break
            
            # Recalculate bounds for new position
            bounds = self._get_shape_bounds(shape, test_x, test_y)
        
        return (test_x, test_y)
    
    def _get_shape_bounds(self, shape, center_x, center_y):
        """Calculate the bounding box of a shape at a given position, including rotation"""
        if isinstance(shape, ScalableRectangle):
            orig_rect = shape.rect()
            width = orig_rect.width()
            height = orig_rect.height()
            rotation = getattr(shape, 'current_rotation', 0)
            
            # For rotated rectangles, we need to calculate the actual bounds
            if rotation != 0:
                # Approximate expanded bounds for rotated rectangle
                diagonal = (width**2 + height**2)**0.5
                half_diagonal = diagonal / 2
                return {
                    'left': center_x - half_diagonal,
                    'right': center_x + half_diagonal,
                    'top': center_y - half_diagonal,
                    'bottom': center_y + half_diagonal
                }
            else:
                return {
                    'left': center_x - width/2,
                    'right': center_x + width/2,
                    'top': center_y - height/2,
                    'bottom': center_y + height/2
                }
        else:  # ScalableTriangle
            orig_size = getattr(shape, 'size', 10)
            rotation = getattr(shape, 'current_rotation', 0)
            
            # For rotated triangles, use expanded bounds
            if rotation != 0:
                # Approximate expanded bounds for rotated triangle
                diagonal = orig_size * 1.414  # sqrt(2) approximation
                half_diagonal = diagonal / 2
                return {
                    'left': center_x - half_diagonal,
                    'right': center_x + half_diagonal,
                    'top': center_y - half_diagonal,
                    'bottom': center_y + half_diagonal
                }
            else:
                return {
                    'left': center_x - orig_size/2,
                    'right': center_x + orig_size/2,
                    'top': center_y - orig_size/2,
                    'bottom': center_y + orig_size/2
                }
    
    def _check_overlap_with_existing(self, bounds):
        """Check if the given bounds overlap with any existing shapes"""
        for existing_bounds in self.placed_ref_shapes:
            if self._bounds_overlap(bounds, existing_bounds):
                return True
        return False
    
    def _bounds_overlap(self, bounds1, bounds2):
        """Check if two bounding boxes overlap"""
        return not (bounds1['right'] <= bounds2['left'] or 
                   bounds1['left'] >= bounds2['right'] or
                   bounds1['bottom'] <= bounds2['top'] or
                   bounds1['top'] >= bounds2['bottom'])
    
    def _draw_single_reference_shape(self, original_shape, original_index, center_x, center_y):
        """Helper method to draw a single reference shape"""
        if isinstance(original_shape, ScalableRectangle):
            # Get original rectangle dimensions from rect()
            orig_rect = original_shape.rect()
            width = orig_rect.width()
            height = orig_rect.height()
            
            # Create rectangle at (0,0) with original dimensions
            ref_shape = QGraphicsRectItem(0, 0, width, height)
            
            # Position the shape so its center is at the target center
            ref_shape.setPos(center_x - width/2, center_y - height/2)
            
            # Apply original rotation
            original_rotation = getattr(original_shape, 'current_rotation', 0)
            if original_rotation != 0:
                # Set rotation center to center of rectangle
                rect_center = QPointF(width/2, height/2)
                ref_shape.setTransformOriginPoint(rect_center)
                ref_shape.setRotation(original_rotation)
                
        else:  # ScalableTriangle
            # Get original triangle size
            orig_size = getattr(original_shape, 'size', 10)  # Default to 10 if no size
            triangle_points = [
                QPointF(0, 0),              # Top-left corner
                QPointF(orig_size, 0),      # Top-right corner 
                QPointF(0, orig_size)       # Bottom-left corner
            ]
            triangle_polygon = QPolygonF(triangle_points)
            ref_shape = QGraphicsPolygonItem(triangle_polygon)
            
            # Position the triangle so its center is at the target center
            ref_shape.setPos(center_x - orig_size/2, center_y - orig_size/2)
            
            # Apply original rotation with geometric center as pivot
            original_rotation = getattr(original_shape, 'current_rotation', 0)
            if original_rotation != 0:
                # Set rotation center to geometric center (centroid)
                triangle_center = QPointF(orig_size/3, orig_size/3)
                ref_shape.setTransformOriginPoint(triangle_center)
                ref_shape.setRotation(original_rotation)
        
        # Get original colors from the shape
        original_fill_color = getattr(original_shape, 'original_fill_color', '')
        original_frame_color = getattr(original_shape, 'original_frame_color', '#8B4513')
        original_is_filled = getattr(original_shape, 'original_is_filled', False)
        
        # Set appearance using original colors with black frame
        ref_shape.setPen(QPen(QColor(0, 0, 0), 0.5))  # Always black frame
        
        # Set fill color - use original fill color if shape was filled, otherwise light gray
        if original_is_filled and original_fill_color:
            fill_color = QColor(original_fill_color)
            ref_shape.setBrush(QBrush(fill_color))
        else:
            # Default to light gray if no fill color or not filled
            ref_shape.setBrush(QBrush(QColor(200, 200, 200)))
        
        # Set high z-value to ensure it appears on top
        ref_shape.setZValue(90)
        
        # Add number label inside the reference shape (use original array index + 1)
        number_text = QGraphicsTextItem(str(original_index + 1))
        number_font = QFont()
        number_font.setPointSize(1)  # Smallest possible font size
        number_font.setBold(True)
        number_text.setFont(number_font)
        number_text.setDefaultTextColor(QColor(0, 0, 0))
        
        # Get text bounding box to properly center it
        text_rect = number_text.boundingRect()
        
        # Calculate text position based on shape type
        if isinstance(original_shape, ScalableTriangle):
            # For triangles, position text at geometric center (centroid)
            # Triangle is positioned at center_x - orig_size/2, center_y - orig_size/2
            # Centroid is at (size/3, size/3) from triangle's top-left corner
            orig_size = getattr(original_shape, 'size', 10)
            triangle_top_left_x = center_x - orig_size/2
            triangle_top_left_y = center_y - orig_size/2
            text_center_x = triangle_top_left_x + orig_size/3
            text_center_y = triangle_top_left_y + orig_size/3
            number_text.setPos(text_center_x - text_rect.width()/2, text_center_y - text_rect.height()/2)
        else:
            # For rectangles, position at the provided center
            number_text.setPos(center_x - text_rect.width()/2, center_y - text_rect.height()/2)
        
        number_text.setZValue(100)
        
        # Add to scene and track for removal
        self.scene.addItem(ref_shape)
        self.scene.addItem(number_text)
        self.number_text_items.append(ref_shape)  # Track for removal
        self.number_text_items.append(number_text)  # Track for removal
    
    def hide_shape_numbers(self):
        """Hide all shape serial numbers"""
        if not self.numbers_visible:
            return  # Already hidden
        
        self.numbers_visible = False
        
        # Remove all number text items
        for text_item in self.number_text_items:
            if text_item.scene():
                self.scene.removeItem(text_item)
        
        self.number_text_items.clear()
        self.placed_ref_shapes.clear()  # Clear collision tracking
        print("Hidden shape numbers")
    
    def create_grid(self):
        """Create a 6x6 grid with boxes of 250x250 pixels each"""
        if self.grid_visible:
            return  # Grid already visible
        
        # Clear any existing grid
        self.clear_grid()
        
        # Grid parameters - 6x6 boxes of 250x250 pixels
        box_size = 250
        grid_cols = 6
        grid_rows = 6
        
        # Starting position (top-left of grid) with offset
        start_x = 0 + self.grid_offset_x
        start_y = 0 + self.grid_offset_y
        
        # Create vertical lines (7 lines to make 6 columns)
        for i in range(grid_cols + 1):
            x = start_x + (i * box_size)
            line_item = self.scene.addLine(
                x, start_y, 
                x, start_y + (grid_rows * box_size),
                QPen(QColor(0, 0, 255), 0)  # Blue pen with minimal thickness
            )
            line_item.setZValue(-0.5)  # Put grid behind shapes but in front of background
            self.grid_items.append(line_item)
        
        # Create horizontal lines (7 lines to make 6 rows)
        for i in range(grid_rows + 1):
            y = start_y + (i * box_size)
            line_item = self.scene.addLine(
                start_x, y,
                start_x + (grid_cols * box_size), y,
                QPen(QColor(0, 0, 255), 0)  # Blue pen with minimal thickness
            )
            line_item.setZValue(-0.5)  # Put grid behind shapes but in front of background
            self.grid_items.append(line_item)
        
        # Create the draggable handle at the top-left corner
        self.grid_handle = GridHandle(self)
        self.grid_handle.setPos(start_x, start_y)
        self.scene.addItem(self.grid_handle)
        
        # Create labels for vertical lines (A, B, C... on top)
        for i in range(grid_cols + 1):
            x = start_x + (i * box_size)
            label_text = chr(ord('A') + i)  # A, B, C, D, E, F, G
            label_item = QGraphicsTextItem(label_text)
            label_item.setPos(x - 10, start_y - 25)  # Position above the grid
            label_item.setDefaultTextColor(QColor(0, 0, 255))  # Blue color to match grid
            label_item.setZValue(-0.4)  # In front of grid lines but behind shapes
            self.scene.addItem(label_item)
            self.grid_labels.append(label_item)
        
        # Create labels for horizontal lines (1, 2, 3... on left)
        for i in range(grid_rows + 1):
            y = start_y + (i * box_size)
            label_text = str(i + 1)  # 1, 2, 3, 4, 5, 6, 7
            label_item = QGraphicsTextItem(label_text)
            label_item.setPos(start_x - 25, y - 10)  # Position to the left of the grid
            label_item.setDefaultTextColor(QColor(0, 0, 255))  # Blue color to match grid
            label_item.setZValue(-0.4)  # In front of grid lines but behind shapes
            self.scene.addItem(label_item)
            self.grid_labels.append(label_item)
        
        self.grid_visible = True
    
    def update_grid_position(self):
        """Update the position of all grid lines and labels based on the handle position"""
        if not self.grid_visible or not self.grid_items:
            return
        
        # Remove existing grid lines
        for item in self.grid_items:
            self.scene.removeItem(item)
        self.grid_items.clear()
        
        # Remove existing labels
        for item in self.grid_labels:
            self.scene.removeItem(item)
        self.grid_labels.clear()
        
        # Grid parameters
        box_size = 250
        grid_cols = 6
        grid_rows = 6
        
        # Starting position with current offset
        start_x = 0 + self.grid_offset_x
        start_y = 0 + self.grid_offset_y
        
        # Recreate vertical lines
        for i in range(grid_cols + 1):
            x = start_x + (i * box_size)
            line_item = self.scene.addLine(
                x, start_y, 
                x, start_y + (grid_rows * box_size),
                QPen(QColor(0, 0, 255), 0)
            )
            line_item.setZValue(-0.5)
            self.grid_items.append(line_item)
        
        # Recreate horizontal lines
        for i in range(grid_rows + 1):
            y = start_y + (i * box_size)
            line_item = self.scene.addLine(
                start_x, y,
                start_x + (grid_cols * box_size), y,
                QPen(QColor(0, 0, 255), 0)
            )
            line_item.setZValue(-0.5)
            self.grid_items.append(line_item)
        
        # Recreate labels for vertical lines (A, B, C... on top)
        for i in range(grid_cols + 1):
            x = start_x + (i * box_size)
            label_text = chr(ord('A') + i)  # A, B, C, D, E, F, G
            label_item = QGraphicsTextItem(label_text)
            label_item.setPos(x - 10, start_y - 25)  # Position above the grid
            label_item.setDefaultTextColor(QColor(0, 0, 255))  # Blue color to match grid
            label_item.setZValue(-0.4)  # In front of grid lines but behind shapes
            self.scene.addItem(label_item)
            self.grid_labels.append(label_item)
        
        # Recreate labels for horizontal lines (1, 2, 3... on left)
        for i in range(grid_rows + 1):
            y = start_y + (i * box_size)
            label_text = str(i + 1)  # 1, 2, 3, 4, 5, 6, 7
            label_item = QGraphicsTextItem(label_text)
            label_item.setPos(start_x - 25, y - 10)  # Position to the left of the grid
            label_item.setDefaultTextColor(QColor(0, 0, 255))  # Blue color to match grid
            label_item.setZValue(-0.4)  # In front of grid lines but behind shapes
            self.scene.addItem(label_item)
            self.grid_labels.append(label_item)
    
    def clear_grid(self):
        """Remove the grid, labels, and handle"""
        for item in self.grid_items:
            self.scene.removeItem(item)
        self.grid_items.clear()
        
        for item in self.grid_labels:
            self.scene.removeItem(item)
        self.grid_labels.clear()
        
        if self.grid_handle:
            self.scene.removeItem(self.grid_handle)
            self.grid_handle = None
        
        self.grid_visible = False

    def fill_A1_and_A2_boxes(self):
        """Fill all boxes that contain shapes with different colors and color the overlapping shapes"""
        if not self.grid_visible:
            return  # No grid to reference
        
        # Clear previous inclusion data
        self.box_inclusion_data = {}
        
        # Calculate box positions and size
        box_size = 250
        grid_cols = 6
        grid_rows = 6
        
        # Define colors for each box (36 different colors for 6x6 grid)
        # Using distinct, high-contrast colors that are easy for OpenCV to detect
        box_colors = [
            QColor(255, 0, 0),      # Bright Red
            QColor(0, 255, 0),      # Bright Green  
            QColor(0, 0, 255),      # Bright Blue
            QColor(255, 255, 0),    # Bright Yellow
            QColor(255, 0, 255),    # Bright Magenta
            QColor(0, 255, 255),    # Bright Cyan
            QColor(255, 128, 0),    # Orange
            QColor(128, 0, 255),    # Purple
            QColor(255, 0, 128),    # Hot Pink
            QColor(0, 128, 255),    # Sky Blue
            QColor(128, 255, 0),    # Lime Green
            QColor(255, 64, 64),    # Light Red
            QColor(64, 255, 64),    # Light Green
            QColor(64, 64, 255),    # Light Blue
            QColor(255, 255, 64),   # Light Yellow
            QColor(255, 64, 255),   # Light Magenta
            QColor(64, 255, 255),   # Light Cyan
            QColor(192, 0, 0),      # Dark Red
            QColor(0, 192, 0),      # Dark Green
            QColor(0, 0, 192),      # Dark Blue
            QColor(192, 192, 0),    # Dark Yellow
            QColor(192, 0, 192),    # Dark Magenta
            QColor(0, 192, 192),    # Dark Cyan
            QColor(255, 96, 0),     # Red Orange
            QColor(255, 0, 96),     # Pink Red
            QColor(96, 255, 0),     # Yellow Green
            QColor(0, 255, 96),     # Green Cyan  
            QColor(96, 0, 255),     # Blue Purple
            QColor(0, 96, 255),     # Cyan Blue
            QColor(255, 192, 0),    # Golden Orange
            QColor(255, 0, 192),    # Magenta Pink
            QColor(192, 255, 0),    # Lime Yellow
            QColor(0, 255, 192),    # Cyan Green
            QColor(192, 0, 255),    # Purple Magenta
            QColor(0, 192, 255),    # Blue Cyan
            QColor(128, 64, 0)      # Brown
        ]
        
        # Find which boxes contain shapes and collect box information
        boxes_with_shapes = []
        
        for row in range(grid_rows):
            for col in range(grid_cols):
                # Calculate box position
                box_x = self.grid_offset_x + (col * box_size)
                box_y = self.grid_offset_y + (row * box_size)
                box_rect = QRectF(box_x, box_y, box_size, box_size)
                
                # Check if any shapes overlap with this box
                has_shapes = False
                for item in self.scene.items():
                    if (item != self.background_item and 
                        item not in self.grid_items and 
                        item not in self.grid_labels and
                        item not in self.cut_lines and
                        item != self.grid_handle and
                        (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                        
                        shape_rect = item.sceneBoundingRect()
                        if box_rect.intersects(shape_rect):
                            has_shapes = True
                            break
                
                if has_shapes:
                    # Calculate box index for color selection based on grid position
                    # This ensures each box always gets the same color regardless of order
                    box_index = row * grid_cols + col  # A1=0, B1=1, C1=2, A2=6, etc.
                    color = box_colors[box_index % len(box_colors)]
                    
                    boxes_with_shapes.append({
                        'rect': box_rect,
                        'color': color,
                        'x': box_x,
                        'y': box_y,
                        'row': row,
                        'col': col,
                        'box_index': box_index  # Store for consistent identification
                    })
        
        # Create colored rectangles for boxes that contain shapes
        for box_info in boxes_with_shapes:
            colored_rect = QGraphicsRectItem(box_info['x'], box_info['y'], box_size, box_size)
            colored_rect.setPen(QPen(Qt.transparent))  # No border
            colored_rect.setBrush(QBrush(box_info['color']))  # Box color
            colored_rect.setZValue(-0.3)
            self.scene.addItem(colored_rect)
            self.cut_lines.append(colored_rect)
        
        # Color shapes based on which box they primarily belong to
        for item in self.scene.items():
            # Only check actual shapes (rectangles and triangles)
            if (item != self.background_item and 
                item not in self.grid_items and 
                item not in self.grid_labels and
                item not in self.cut_lines and
                item != self.grid_handle and
                (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                
                # Get the shape's bounding rectangle in scene coordinates
                shape_rect = item.sceneBoundingRect()
                
                # Find which box has the largest overlap with this shape
                max_overlap_area = 0
                best_box_color = None
                best_box_info = None
                
                for box_info in boxes_with_shapes:
                    if box_info['rect'].intersects(shape_rect):
                        intersection_rect = box_info['rect'].intersected(shape_rect)
                        overlap_area = intersection_rect.width() * intersection_rect.height()
                        
                        if overlap_area > max_overlap_area:
                            max_overlap_area = overlap_area
                            best_box_color = box_info['color']
                            best_box_info = box_info
                
                # Color the shape based on overlap
                if best_box_color and max_overlap_area > 0:
                    total_shape_area = shape_rect.width() * shape_rect.height()
                    area_ratio = max_overlap_area / total_shape_area if total_shape_area > 0 else 0
                    
                    if area_ratio > 0.25:  # More than 25% of shape is in the dominant box (lowered threshold)
                        # Fill with the box color
                        item.setBrush(QBrush(best_box_color))  # Box color
                        item.setPen(QPen(best_box_color, 0))  # Matching frame
                        
                        # Store inclusion data - this shape belongs to this box
                        box_index = best_box_info['box_index']
                        if box_index not in self.box_inclusion_data:
                            self.box_inclusion_data[box_index] = []
                        self.box_inclusion_data[box_index].append(item)
                        
                    else:  # Less than 25% of shape is in any box
                        # Fill with white
                        item.setBrush(QBrush(QColor(255, 255, 255)))  # Solid white
                        item.setPen(QPen(QColor(0, 0, 0), 0))  # Black frame
    
    def fill_all_boxes_white(self):
        """Fill all boxes with white color after saving files"""
        if not self.grid_visible:
            return  # No grid to reference
        
        # Calculate box positions and size
        box_size = 250
        grid_cols = 6
        grid_rows = 6
        
        print("Filling all boxes with white color...")
        
        # Clear existing colored boxes first
        items_to_remove = []
        for cut_item in self.cut_lines:
            if isinstance(cut_item, QGraphicsRectItem) and cut_item.brush().color() != Qt.transparent:
                self.scene.removeItem(cut_item)
                items_to_remove.append(cut_item)
        
        # Remove them from cut_lines list
        for item in items_to_remove:
            self.cut_lines.remove(item)
        
        # Create white rectangles for all boxes
        for row in range(grid_rows):
            for col in range(grid_cols):
                # Calculate box position
                box_x = self.grid_offset_x + (col * box_size)
                box_y = self.grid_offset_y + (row * box_size)
                
                # Create white filled rectangle for this box
                white_rect = QGraphicsRectItem(box_x, box_y, box_size, box_size)
                white_rect.setPen(QPen(Qt.transparent))  # No border
                white_rect.setBrush(QBrush(QColor(255, 255, 255)))  # White fill
                white_rect.setZValue(-0.3)
                self.scene.addItem(white_rect)
                self.cut_lines.append(white_rect)
        
        print(f"Filled all {grid_rows * grid_cols} boxes with white color")

    def draw_red_green_border(self):
        """Detect all colored blobs and show their borders"""
        try:
            if cv2 is None:
                print("OpenCV not available, blob detection disabled")
                return
            
            # Get the scene bounds - focus on the grid area only
            if not self.grid_visible:
                print("Grid not visible, cannot detect colored areas")
                return
                
            # Calculate the actual grid area
            box_size = 250
            grid_cols = 6
            grid_rows = 6
            
            # Define the grid area bounds
            grid_left = self.grid_offset_x
            grid_top = self.grid_offset_y
            grid_right = grid_left + (grid_cols * box_size)
            grid_bottom = grid_top + (grid_rows * box_size)
            
            grid_rect = QRectF(grid_left, grid_top, grid_right - grid_left, grid_bottom - grid_top)
            
            print(f"Detecting colored blobs in grid area: {grid_rect}")
            
            # Create an image representation of the grid area only
            scene_image = self.render_scene_to_image(grid_rect)
            if scene_image is None:
                print("Failed to render scene to image")
                return
            
            h, w, ch = scene_image.shape
            print(f"Rendered image size: {w}x{h}")
            
            # Create precise color detection based on the exact box colors used
            # Convert QColor RGB values to BGR ranges for OpenCV (with small tolerance)
            # Using distinct, high-contrast colors that are easy for OpenCV to detect
            box_colors_rgb = [
                (255, 0, 0),        # Bright Red
                (0, 255, 0),        # Bright Green  
                (0, 0, 255),        # Bright Blue
                (255, 255, 0),      # Bright Yellow
                (255, 0, 255),      # Bright Magenta
                (0, 255, 255),      # Bright Cyan
                (255, 128, 0),      # Orange
                (128, 0, 255),      # Purple
                (255, 0, 128),      # Hot Pink
                (0, 128, 255),      # Sky Blue
                (128, 255, 0),      # Lime Green
                (255, 64, 64),      # Light Red
                (64, 255, 64),      # Light Green
                (64, 64, 255),      # Light Blue
                (255, 255, 64),     # Light Yellow
                (255, 64, 255),     # Light Magenta
                (64, 255, 255),     # Light Cyan
                (192, 0, 0),        # Dark Red
                (0, 192, 0),        # Dark Green
                (0, 0, 192),        # Dark Blue
                (192, 192, 0),      # Dark Yellow
                (192, 0, 192),      # Dark Magenta
                (0, 192, 192),      # Dark Cyan
                (255, 96, 0),       # Red Orange
                (255, 0, 96),       # Pink Red
                (96, 255, 0),       # Yellow Green
                (0, 255, 96),       # Green Cyan  
                (96, 0, 255),       # Blue Purple
                (0, 96, 255),       # Cyan Blue
                (255, 192, 0),      # Golden Orange
                (255, 0, 192),      # Magenta Pink
                (192, 255, 0),      # Lime Yellow
                (0, 255, 192),      # Cyan Green
                (192, 0, 255),      # Purple Magenta
                (0, 192, 255),      # Blue Cyan
                (128, 64, 0)        # Brown
            ]
            
            # Create color detection with precise ranges for each box color
            tolerance = 15  # Small tolerance for color matching
            colors = {}
            
            for i, (r, g, b) in enumerate(box_colors_rgb):
                # Convert RGB to BGR for OpenCV
                bgr_color = (b, g, r)
                color_name = f"color_{i:02d}"
                
                colors[color_name] = {
                    'lower': np.array([max(0, bgr_color[0] - tolerance), 
                                     max(0, bgr_color[1] - tolerance), 
                                     max(0, bgr_color[2] - tolerance)]),
                    'upper': np.array([min(255, bgr_color[0] + tolerance), 
                                     min(255, bgr_color[1] + tolerance), 
                                     min(255, bgr_color[2] + tolerance)]),
                    'border_color': QColor(0, 0, 0)  # Black border
                }
            
            borders_created = 0
            
            # Detect blobs for each specific color
            for color_name, color_info in colors.items():
                # Create mask for this specific color
                mask = cv2.inRange(scene_image, color_info['lower'], color_info['upper'])
                
                # Clean up the mask
                kernel = np.ones((3, 3), np.uint8)
                mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
                mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
                
                # Find contours for this color
                contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                
                # Draw borders for each blob of this color
                for contour in contours:
                    if cv2.contourArea(contour) < 100:  # Skip very small areas
                        continue
                    
                    # Convert contour points to scene coordinates and create polygon
                    polygon_points = []
                    for point in contour:
                        x, y = point[0]
                        scene_x = grid_left + x
                        scene_y = grid_top + y
                        polygon_points.append(QPointF(scene_x, scene_y))
                    
                    if len(polygon_points) < 3:
                        continue
                    
                    # Create polygon item for the blob border
                    polygon = QPolygonF(polygon_points)
                    polygon_item = QGraphicsPolygonItem(polygon)
                    
                    # Set border style - thin black border, no fill
                    border_pen = QPen(color_info['border_color'], 1)  # 1 pixel thin black border
                    border_pen.setCosmetic(True)
                    polygon_item.setPen(border_pen)
                    polygon_item.setBrush(QBrush(Qt.transparent))  # No fill
                    polygon_item.setZValue(2)  # Put borders in front of everything
                    
                    # Add to scene and track as cut line
                    self.scene.addItem(polygon_item)
                    self.cut_lines.append(polygon_item)
                    borders_created += 1
                    
                    # Find which grid box this blob belongs to
                    blob_center_x = sum(pt.x() for pt in polygon_points) / len(polygon_points)
                    blob_center_y = sum(pt.y() for pt in polygon_points) / len(polygon_points)
                    
                    # Calculate which box this blob is in
                    box_col = int((blob_center_x - grid_left) // box_size)
                    box_row = int((blob_center_y - grid_top) // box_size)
                    
                    # Make sure we're within the grid bounds
                    if 0 <= box_col < grid_cols and 0 <= box_row < grid_rows:
                        # Calculate the top-left corner of this box
                        box_left = grid_left + (box_col * box_size)
                        box_top = grid_top + (box_row * box_size)
                        
                        # Calculate circle positions relative to box top-left corner
                        # Circle 1: half width, quarter height (top of triangle)
                        circle1_x = box_left + (box_size / 2)
                        circle1_y = box_top + (box_size / 4)
                        
                        # Circle 2: quarter width, three quarters height (bottom left)
                        circle2_x = box_left + (box_size / 4)
                        circle2_y = box_top + (3 * box_size / 4)
                        
                        # Circle 3: three quarters width, three quarters height (bottom right)
                        circle3_x = box_left + (3 * box_size / 4)
                        circle3_y = box_top + (3 * box_size / 4)
                        
                        # Create circles with radius 3 (half the original radius)
                        circle_radius = 3
                        
                        # Circle 1 (top)
                        circle1 = self.scene.addEllipse(
                            circle1_x - circle_radius, circle1_y - circle_radius,
                            circle_radius * 2, circle_radius * 2,
                            QPen(QColor(0, 0, 0), 1),  # Black border
                            QBrush(Qt.transparent)     # No fill
                        )
                        circle1.setZValue(3)  # In front of borders
                        self.cut_lines.append(circle1)
                        
                        # Circle 2 (bottom left)
                        circle2 = self.scene.addEllipse(
                            circle2_x - circle_radius, circle2_y - circle_radius,
                            circle_radius * 2, circle_radius * 2,
                            QPen(QColor(0, 0, 0), 1),  # Black border
                            QBrush(Qt.transparent)     # No fill
                        )
                        circle2.setZValue(3)  # In front of borders
                        self.cut_lines.append(circle2)
                        
                        # Circle 3 (bottom right)
                        circle3 = self.scene.addEllipse(
                            circle3_x - circle_radius, circle3_y - circle_radius,
                            circle_radius * 2, circle_radius * 2,
                            QPen(QColor(0, 0, 0), 1),  # Black border
                            QBrush(Qt.transparent)     # No fill
                        )
                        circle3.setZValue(3)  # In front of borders
                        self.cut_lines.append(circle3)
                        
                        # Add line-drawn text label on screen for visual confirmation
                        col_letter = chr(ord('A') + box_col)
                        row_number = box_row + 1
                        box_name = f"{col_letter}{row_number}"
                        
                        # Draw the box name using lines instead of text
                        self.draw_line_text(box_name, circle1_x - 10, circle1_y - 35)
                        
                        # Create SVG file for this blob
                        self.create_blob_svg(polygon_points, circle1_x, circle1_y, circle2_x, circle2_y, circle3_x, circle3_y,
                                           circle_radius, box_row, box_col)
                        
                        # Create DXF file for this blob
                        self.create_blob_dxf(polygon_points, circle1_x, circle1_y, circle2_x, circle2_y, circle3_x, circle3_y,
                                           circle_radius, box_row, box_col)
            
            print(f"Created {borders_created} blob borders using precise color detection")
            
            # Draw thin black frames around all shapes
            self.draw_shape_frames()
                
        except Exception as e:
            print(f"Error in blob detection and border drawing: {e}")
    
    def create_blob_svg(self, polygon_points, circle1_x, circle1_y, circle2_x, circle2_y, circle3_x, circle3_y, circle_radius, box_row, box_col):
        """Create an SVG file for a single blob with its border and three circles"""
        try:
            # Create blobs directory if it doesn't exist
            blobs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blobs")
            if not os.path.exists(blobs_dir):
                os.makedirs(blobs_dir)
            
            # Calculate box name (A1, B2, etc.)
            col_letter = chr(ord('A') + box_col)
            row_number = box_row + 1
            box_name = f"{col_letter}{row_number}"
            
            # Calculate bounding box of the polygon
            min_x = min(pt.x() for pt in polygon_points)
            max_x = max(pt.x() for pt in polygon_points)
            min_y = min(pt.y() for pt in polygon_points)
            max_y = max(pt.y() for pt in polygon_points)
            
            # Add some padding
            padding = 10
            svg_min_x = min_x - padding
            svg_min_y = min_y - padding
            svg_width = (max_x - min_x) + (2 * padding)
            svg_height = (max_y - min_y) + (2 * padding)
            
            # Create SVG content
            svg_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" 
     width="{svg_width:.2f}" height="{svg_height:.2f}" 
     viewBox="{svg_min_x:.2f} {svg_min_y:.2f} {svg_width:.2f} {svg_height:.2f}">
  
  <!-- Blob border -->
  <polygon points="'''
            
            # Add polygon points
            point_strings = []
            for pt in polygon_points:
                point_strings.append(f"{pt.x():.2f},{pt.y():.2f}")
            svg_content += " ".join(point_strings)
            
            svg_content += f'''" 
           fill="none" 
           stroke="black" 
           stroke-width="1"/>
  
  <!-- Circle 1 (top) -->
  <circle cx="{circle1_x:.2f}" cy="{circle1_y:.2f}" r="{circle_radius}" 
          fill="none" stroke="black" stroke-width="1"/>
  
  <!-- Circle 2 (bottom left) -->
  <circle cx="{circle2_x:.2f}" cy="{circle2_y:.2f}" r="{circle_radius}" 
          fill="none" stroke="black" stroke-width="1"/>
  
  <!-- Circle 3 (bottom right) -->
  <circle cx="{circle3_x:.2f}" cy="{circle3_y:.2f}" r="{circle_radius}" 
          fill="none" stroke="black" stroke-width="1"/>
  
  <!-- Box name label using lines -->'''
            
            # Add line-drawn text for the box name
            text_x = circle1_x - 10
            text_y = circle1_y - 25
            
            svg_content += self.get_svg_line_text(box_name, text_x, text_y)
            
            svg_content += '''
        
</svg>'''
            
            # Save SVG file
            svg_filename = f"{box_name}_blob.svg"
            svg_path = os.path.join(blobs_dir, svg_filename)
            
            with open(svg_path, 'w', encoding='utf-8') as svg_file:
                svg_file.write(svg_content)
            
            print(f"Created SVG file: {svg_path}")
            
        except Exception as e:
            print(f"Error creating SVG for box {box_col},{box_row}: {e}")
    
    def get_svg_line_text(self, text, start_x, start_y):
        """Generate SVG line elements for text characters"""
        try:
            svg_lines = []
            char_width = 8
            char_height = 12
            x_offset = 0
            
            for char in text:
                char_x = start_x + x_offset
                lines = self.get_character_lines(char, char_x, start_y, char_width, char_height)
                
                for line in lines:
                    x1, y1, x2, y2 = line
                    svg_lines.append(f'  <line x1="{x1:.2f}" y1="{y1:.2f}" x2="{x2:.2f}" y2="{y2:.2f}" stroke="black" stroke-width="1"/>')
                
                x_offset += char_width + 2
            
            return '\n'.join(svg_lines)
            
        except Exception as e:
            print(f"Error generating SVG line text: {e}")
            return ""
    
    def get_character_lines(self, char, char_x, start_y, char_width, char_height):
        """Get line definitions for a character"""
        lines = []
        
        if char == 'A':
            lines = [
                (char_x, start_y + char_height, char_x, start_y + 2),
                (char_x + char_width, start_y + char_height, char_x + char_width, start_y + 2),
                (char_x, start_y + 2, char_x + char_width - 2, start_y + 2),  # Gap for laser cutting
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2)
            ]
        elif char == 'B':
            lines = [
                (char_x, start_y, char_x, start_y + char_height),
                (char_x, start_y, char_x + char_width - 2, start_y),  # Gap for laser cutting
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                (char_x, start_y + char_height, char_x + char_width - 2, start_y + char_height),  # Gap for laser cutting
                (char_x + char_width, start_y + 2, char_x + char_width, start_y + char_height//2),  # Gap for laser cutting
                (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height - 2)  # Gap for laser cutting
            ]
        elif char == 'C':
            lines = [
                (char_x, start_y + 2, char_x, start_y + char_height - 2),
                (char_x, start_y + 2, char_x + char_width, start_y + 2),
                (char_x, start_y + char_height - 2, char_x + char_width, start_y + char_height - 2)
            ]
        elif char == 'D':
            lines = [
                # Left vertical line
                (char_x, start_y, char_x, start_y + char_height),
                # Top horizontal line (with gap for laser cutting)
                (char_x, start_y, char_x + char_width - 2, start_y),
                # Bottom horizontal line (with gap for laser cutting)
                (char_x, start_y + char_height, char_x + char_width - 2, start_y + char_height),
                # Right curve (approximated with lines, with gaps)
                (char_x + char_width - 2, start_y, char_x + char_width, start_y + 2),
                (char_x + char_width, start_y + 2, char_x + char_width, start_y + char_height - 2),
                (char_x + char_width, start_y + char_height - 2, char_x + char_width - 2, start_y + char_height)
            ]
        elif char == '1':
            lines = [
                (char_x + char_width//2, start_y, char_x + char_width//2, start_y + char_height),
                (char_x + char_width//4, start_y + 2, char_x + char_width//2, start_y),
                (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
            ]
        elif char == '2':
            lines = [
                (char_x, start_y, char_x + char_width, start_y),
                (char_x + char_width, start_y, char_x + char_width, start_y + char_height//2),
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                (char_x, start_y + char_height//2, char_x, start_y + char_height),
                (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
            ]
        elif char == '3':
            lines = [
                (char_x, start_y, char_x + char_width, start_y),
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                (char_x, start_y + char_height, char_x + char_width, start_y + char_height),
                (char_x + char_width, start_y, char_x + char_width, start_y + char_height)
            ]
        elif char == '4':
            lines = [
                (char_x, start_y, char_x, start_y + char_height//2),
                (char_x + char_width, start_y, char_x + char_width, start_y + char_height),
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2)
            ]
        elif char == '5':
            lines = [
                (char_x, start_y, char_x + char_width, start_y),
                (char_x, start_y, char_x, start_y + char_height//2),
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height),
                (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
            ]
        elif char == '6':
            lines = [
                (char_x, start_y, char_x, start_y + char_height),
                (char_x, start_y, char_x + char_width, start_y),
                (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                (char_x, start_y + char_height, char_x + char_width, start_y + char_height),
                (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height)
            ]
        
        return lines
    
    def create_blob_dxf(self, polygon_points, circle1_x, circle1_y, circle2_x, circle2_y, circle3_x, circle3_y, circle_radius, box_row, box_col):
        """Create a DXF file for a single blob with its border and three circles"""
        try:
            if ezdxf is None:
                print("ezdxf not available, skipping DXF creation")
                return
            
            # Create blobs directory if it doesn't exist
            blobs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blobs")
            if not os.path.exists(blobs_dir):
                os.makedirs(blobs_dir)
            
            # Calculate box name (A1, B2, etc.)
            col_letter = chr(ord('A') + box_col)
            row_number = box_row + 1
            box_name = f"{col_letter}{row_number}"
            
            # Create a new DXF document
            doc = ezdxf.new('R2010')  # DXF version R2010 (AutoCAD 2010)
            msp = doc.modelspace()
            
            # Create layers for different elements
            doc.layers.new('BLOB_BORDER', dxfattribs={'color': 1})  # Red
            doc.layers.new('CIRCLES', dxfattribs={'color': 2})      # Yellow
            doc.layers.new('TEXT', dxfattribs={'color': 3})         # Green
            
            # Add blob border as polyline
            if len(polygon_points) >= 3:
                # Convert QPointF to tuples for ezdxf
                dxf_points = [(pt.x(), pt.y()) for pt in polygon_points]
                
                # Create closed polyline
                polyline = msp.add_lwpolyline(dxf_points, close=True)
                polyline.dxf.layer = 'BLOB_BORDER'
            
            # Add circles
            circle1 = msp.add_circle((circle1_x, circle1_y), circle_radius)
            circle1.dxf.layer = 'CIRCLES'
            
            circle2 = msp.add_circle((circle2_x, circle2_y), circle_radius)
            circle2.dxf.layer = 'CIRCLES'
            
            circle3 = msp.add_circle((circle3_x, circle3_y), circle_radius)
            circle3.dxf.layer = 'CIRCLES'
            
            # Add line-drawn text label instead of text entities
            text_x = circle1_x - 10
            text_y = circle1_y - 25
            
            # Draw the box name using lines
            char_width = 8
            char_height = 12
            x_offset = 0
            
            for char in box_name:
                char_x = text_x + x_offset
                lines = self.get_character_lines(char, char_x, text_y, char_width, char_height)
                
                # Add each line to the DXF
                for line in lines:
                    x1, y1, x2, y2 = line
                    line_entity = msp.add_line((x1, y1), (x2, y2))
                    line_entity.dxf.layer = 'TEXT'
                
                x_offset += char_width + 2
            
            # Save DXF file
            dxf_filename = f"{box_name}_blob.dxf"
            dxf_path = os.path.join(blobs_dir, dxf_filename)
            
            doc.saveas(dxf_path)
            print(f"Created DXF file: {dxf_path}")
            
        except Exception as e:
            print(f"Error creating DXF for box {box_col},{box_row}: {e}")
    
    def draw_shape_frames(self):
        """Draw thin black frames around all shapes for better visualization"""
        try:
            frames_created = 0
            
            # Find all shapes (rectangles and triangles) in the scene
            for item in self.scene.items():
                if (item != self.background_item and 
                    item not in self.grid_items and 
                    item not in self.grid_labels and
                    item not in self.cut_lines and
                    item != self.grid_handle and
                    (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                    
                    if isinstance(item, ScalableRectangle):
                        # For rectangles, create a frame that matches exactly including rotation
                        frame_rect = QGraphicsRectItem(item.rect())
                        frame_rect.setPos(item.pos())
                        frame_rect.setRotation(item.rotation())
                        frame_rect.setTransformOriginPoint(item.transformOriginPoint())
                        
                        # Use cosmetic pen for constant thin line regardless of zoom
                        pen = QPen(QColor(0, 0, 0), 0)  # Width 0 = cosmetic (always 1 pixel)
                        pen.setCosmetic(True)
                        frame_rect.setPen(pen)
                        frame_rect.setBrush(QBrush(Qt.transparent))  # No fill
                        frame_rect.setZValue(1.5)  # In front of shapes but behind blob borders
                        
                        self.scene.addItem(frame_rect)
                        self.cut_lines.append(frame_rect)
                        frames_created += 1
                        
                    elif isinstance(item, ScalableTriangle):
                        # For triangles, create a frame that matches exactly including rotation
                        frame_polygon = QGraphicsPolygonItem(item.polygon())
                        frame_polygon.setPos(item.pos())
                        frame_polygon.setRotation(item.rotation())
                        frame_polygon.setTransformOriginPoint(item.transformOriginPoint())
                        
                        # Use cosmetic pen for constant thin line regardless of zoom
                        pen = QPen(QColor(0, 0, 0), 0)  # Width 0 = cosmetic (always 1 pixel)
                        pen.setCosmetic(True)
                        frame_polygon.setPen(pen)
                        frame_polygon.setBrush(QBrush(Qt.transparent))  # No fill
                        frame_polygon.setZValue(1.5)  # In front of shapes but behind blob borders
                        
                        self.scene.addItem(frame_polygon)
                        self.cut_lines.append(frame_polygon)
                        frames_created += 1
            
            print(f"Created {frames_created} thin black frames around shapes")
            
        except Exception as e:
            print(f"Error drawing shape frames: {e}")
    
    def draw_line_text(self, text, start_x, start_y):
        """Draw text using line segments for better DXF compatibility"""
        try:
            char_width = 8
            char_height = 12
            line_width = 1
            
            x_offset = 0
            
            for char in text:
                char_x = start_x + x_offset
                
                # Define line patterns for each character
                lines = []
                
                if char == 'A':
                    lines = [
                        # Left vertical line
                        (char_x, start_y + char_height, char_x, start_y + 2),
                        # Right vertical line  
                        (char_x + char_width, start_y + char_height, char_x + char_width, start_y + 2),
                        # Top horizontal line (with gap for laser cutting)
                        (char_x, start_y + 2, char_x + char_width - 2, start_y + 2),
                        # Middle horizontal line
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2)
                    ]
                elif char == 'B':
                    lines = [
                        # Left vertical line
                        (char_x, start_y, char_x, start_y + char_height),
                        # Top horizontal line (with gap)
                        (char_x, start_y, char_x + char_width - 2, start_y),
                        # Middle horizontal line
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                        # Bottom horizontal line (with gap)
                        (char_x, start_y + char_height, char_x + char_width - 2, start_y + char_height),
                        # Top right vertical (with gap)
                        (char_x + char_width, start_y + 2, char_x + char_width, start_y + char_height//2),
                        # Bottom right vertical (with gap) 
                        (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height - 2)
                    ]
                elif char == 'C':
                    lines = [
                        # Left vertical line
                        (char_x, start_y + 2, char_x, start_y + char_height - 2),
                        # Top horizontal line
                        (char_x, start_y + 2, char_x + char_width, start_y + 2),
                        # Bottom horizontal line
                        (char_x, start_y + char_height - 2, char_x + char_width, start_y + char_height - 2)
                    ]
                elif char == 'D':
                    lines = [
                        # Left vertical line
                        (char_x, start_y, char_x, start_y + char_height),
                        # Top horizontal line (with gap)
                        (char_x, start_y, char_x + char_width - 2, start_y),
                        # Bottom horizontal line (with gap)
                        (char_x, start_y + char_height, char_x + char_width - 2, start_y + char_height),
                        # Right curve (approximated with lines, with gaps)
                        (char_x + char_width - 2, start_y, char_x + char_width, start_y + 2),
                        (char_x + char_width, start_y + 2, char_x + char_width, start_y + char_height - 2),
                        (char_x + char_width, start_y + char_height - 2, char_x + char_width - 2, start_y + char_height)
                    ]
                elif char == 'E':
                    lines = [
                        # Left vertical line
                        (char_x, start_y, char_x, start_y + char_height),
                        # Top horizontal line
                        (char_x, start_y, char_x + char_width, start_y),
                        # Middle horizontal line
                        (char_x, start_y + char_height//2, char_x + char_width//2, start_y + char_height//2),
                        # Bottom horizontal line
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
                    ]
                elif char == 'F':
                    lines = [
                        # Left vertical line
                        (char_x, start_y, char_x, start_y + char_height),
                        # Top horizontal line
                        (char_x, start_y, char_x + char_width, start_y),
                        # Middle horizontal line
                        (char_x, start_y + char_height//2, char_x + char_width//2, start_y + char_height//2)
                    ]
                elif char == 'G':
                    lines = [
                        # Left vertical line
                        (char_x, start_y + 2, char_x, start_y + char_height - 2),
                        # Top horizontal line
                        (char_x, start_y + 2, char_x + char_width, start_y + 2),
                        # Bottom horizontal line
                        (char_x, start_y + char_height - 2, char_x + char_width, start_y + char_height - 2),
                        # Right vertical (bottom half)
                        (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height - 2),
                        # Middle horizontal (right half)
                        (char_x + char_width//2, start_y + char_height//2, char_x + char_width, start_y + char_height//2)
                    ]
                elif char == '1':
                    lines = [
                        # Main vertical line
                        (char_x + char_width//2, start_y, char_x + char_width//2, start_y + char_height),
                        # Top diagonal
                        (char_x + char_width//4, start_y + 2, char_x + char_width//2, start_y),
                        # Bottom horizontal
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
                    ]
                elif char == '2':
                    lines = [
                        # Top horizontal
                        (char_x, start_y, char_x + char_width, start_y),
                        # Top right vertical
                        (char_x + char_width, start_y, char_x + char_width, start_y + char_height//2),
                        # Middle horizontal
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                        # Bottom left vertical
                        (char_x, start_y + char_height//2, char_x, start_y + char_height),
                        # Bottom horizontal
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
                    ]
                elif char == '3':
                    lines = [
                        # Top horizontal
                        (char_x, start_y, char_x + char_width, start_y),
                        # Middle horizontal
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                        # Bottom horizontal
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height),
                        # Right vertical
                        (char_x + char_width, start_y, char_x + char_width, start_y + char_height)
                    ]
                elif char == '4':
                    lines = [
                        # Left vertical (top half)
                        (char_x, start_y, char_x, start_y + char_height//2),
                        # Right vertical (full)
                        (char_x + char_width, start_y, char_x + char_width, start_y + char_height),
                        # Middle horizontal
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2)
                    ]
                elif char == '5':
                    lines = [
                        # Top horizontal
                        (char_x, start_y, char_x + char_width, start_y),
                        # Left vertical (top half)
                        (char_x, start_y, char_x, start_y + char_height//2),
                        # Middle horizontal
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                        # Right vertical (bottom half)
                        (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height),
                        # Bottom horizontal
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height)
                    ]
                elif char == '6':
                    lines = [
                        # Left vertical
                        (char_x, start_y, char_x, start_y + char_height),
                        # Top horizontal
                        (char_x, start_y, char_x + char_width, start_y),
                        # Middle horizontal
                        (char_x, start_y + char_height//2, char_x + char_width, start_y + char_height//2),
                        # Bottom horizontal
                        (char_x, start_y + char_height, char_x + char_width, start_y + char_height),
                        # Right vertical (bottom half)
                        (char_x + char_width, start_y + char_height//2, char_x + char_width, start_y + char_height)
                    ]
                
                # Draw all lines for this character
                for line in lines:
                    x1, y1, x2, y2 = line
                    line_item = self.scene.addLine(x1, y1, x2, y2, QPen(QColor(0, 0, 0), line_width))
                    line_item.setZValue(4)  # In front of everything
                    self.cut_lines.append(line_item)
                
                # Move to next character position
                x_offset += char_width + 2
                
        except Exception as e:
            print(f"Error drawing line text: {e}")
    
    def merge_border_points(self, border_points):
        """Merge border points into longer continuous lines"""
        try:
            if not border_points:
                return []
            
            # Separate vertical and horizontal lines
            vertical_lines = [(x1, y1, x2, y2) for x1, y1, x2, y2, line_type in border_points if line_type == 'vertical']
            horizontal_lines = [(x1, y1, x2, y2) for x1, y1, x2, y2, line_type in border_points if line_type == 'horizontal']
            
            merged_lines = []
            
            # Merge vertical lines
            for line in vertical_lines:
                x1, y1, x2, y2 = line
                merged = False
                for i, merged_line in enumerate(merged_lines):
                    mx1, my1, mx2, my2 = merged_line
                    # Check if this line can extend an existing vertical line
                    if (abs(mx1 - x1) < 2 and abs(mx2 - x2) < 2 and  # Same X position
                        abs(my2 - y1) < 2):  # Adjacent Y position
                        # Extend the existing line
                        merged_lines[i] = (mx1, my1, mx2, y2)
                        merged = True
                        break
                
                if not merged:
                    merged_lines.append(line)
            
            # Merge horizontal lines
            for line in horizontal_lines:
                x1, y1, x2, y2 = line
                merged = False
                for i, merged_line in enumerate(merged_lines):
                    mx1, my1, mx2, my2 = merged_line
                    # Check if this line can extend an existing horizontal line
                    if (abs(my1 - y1) < 2 and abs(my2 - y2) < 2 and  # Same Y position
                        abs(mx2 - x1) < 2):  # Adjacent X position
                        # Extend the existing line
                        merged_lines[i] = (mx1, my1, x2, my2)
                        merged = True
                        break
                
                if not merged:
                    merged_lines.append(line)
            
            return merged_lines
            
        except Exception as e:
            print(f"Error merging border points: {e}")
            return [(x1, y1, x2, y2) for x1, y1, x2, y2, _ in border_points]
    
    def render_scene_to_image(self, scene_rect):
        """Render the scene to a numpy array for OpenCV processing"""
        try:
            width = int(scene_rect.width())
            height = int(scene_rect.height())
            
            # Create a QPixmap to render the scene
            pixmap = QPixmap(width, height)
            pixmap.fill(Qt.white)
            
            # Render the scene to the pixmap
            painter = QPainter(pixmap)
            self.scene.render(painter, QRectF(0, 0, width, height), scene_rect)
            painter.end()
            
            # Convert QPixmap to QImage
            qimage = pixmap.toImage()
            qimage = qimage.convertToFormat(qimage.Format_RGB888)
            
            # Convert QImage to numpy array
            width = qimage.width()
            height = qimage.height()
            ptr = qimage.constBits()
            ptr.setsize(qimage.byteCount())
            arr = np.array(ptr).reshape(height, width, 3)
            
            # Convert from RGB to BGR for OpenCV
            arr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
            
            return arr
            
        except Exception as e:
            print(f"Error rendering scene to image: {e}")
            return None
    
    def get_pixel_color(self, x, y):
        """Get the color at a specific pixel coordinate by checking which item is at that point"""
        try:
            # Create a small rectangle around the point for more accurate detection
            search_rect = QRectF(x - 0.5, y - 0.5, 1, 1)
            items_at_point = self.scene.items(search_rect)
            
            # Look for colored shapes first (highest priority)
            for item in items_at_point:
                if (isinstance(item, (ScalableRectangle, ScalableTriangle)) and
                    item.brush().color() != Qt.transparent):
                    return item.brush().color()
            
            # Then check colored rectangles from cut operation (background color)
            for item in items_at_point:
                if (isinstance(item, QGraphicsRectItem) and 
                    item in self.cut_lines and
                    item.brush().color() != Qt.transparent):
                    return item.brush().color()
            
            # Return transparent if no colored item found
            return QColor(Qt.transparent)
            
        except Exception as e:
            print(f"Error getting pixel color at ({x}, {y}): {e}")
            return QColor(Qt.transparent)
    
    def clear_cut_lines(self):
        """Remove all cut lines and filled boxes, and reset shape colors"""
        for cut_item in self.cut_lines:
            self.scene.removeItem(cut_item)
        self.cut_lines.clear()
        
        # Reset all shape colors back to transparent
        for item in self.scene.items():
            if (item != self.background_item and 
                item not in self.grid_items and 
                item not in self.grid_labels and
                item != self.grid_handle and
                (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                # Reset to transparent fill and black frame
                item.setBrush(QBrush(Qt.transparent))
                item.setPen(QPen(QColor(0, 0, 0), 0))  # Reset to black frame
    
    def update_scale_bars(self):
        """Update the scale bars based on current view state"""
        if not hasattr(self, 'horizontal_scale_bar') or not hasattr(self, 'vertical_scale_bar'):
            return
        
        # Check if scene is still valid
        if not hasattr(self, 'scene') or self.scene is None:
            return
            
        try:
            # Get current transformation matrix
            transform = self.transform()
            scale_factor = transform.m11()  # Horizontal scale factor
            
            # Get visible scene rectangle
            visible_scene_rect = self.mapToScene(self.viewport().rect()).boundingRect()
            
            # Update scale bars
            self.horizontal_scale_bar.update_scale(scale_factor, self.scene.sceneRect(), visible_scene_rect)
            self.vertical_scale_bar.update_scale(scale_factor, self.scene.sceneRect(), visible_scene_rect)
        except RuntimeError:
            # Scene has been deleted, skip update
            return
    
    def resizeEvent(self, event):
        """Handle resize events to reposition scale bars"""
        super().resizeEvent(event)
        
        if hasattr(self, 'horizontal_scale_bar') and hasattr(self, 'vertical_scale_bar'):
            # Position horizontal scale bar at top, accounting for the full widget width
            self.horizontal_scale_bar.setGeometry(30, 0, self.width() - 30, 30)
            
            # Position vertical scale bar at left, accounting for the full widget height
            self.vertical_scale_bar.setGeometry(0, 30, 30, self.height() - 30)
            
            # Update scale bars
            self.update_scale_bars()
    
    def showEvent(self, event):
        """Handle show events to initialize scale bars"""
        super().showEvent(event)
        QTimer.singleShot(50, self.update_scale_bars)

class CutterWindow(QMainWindow):
    """Main window for the cutter application"""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cutter - Shape Viewer")
        self.setGeometry(100, 100, 1200, 800)
        
        # Store the path of the currently imported CSV file
        self.current_csv_file = None
        
        # Store DXF items separately for independent clearing
        self.dxf_items = []
        
        # Toggle for overlap detection in organize mode
        self.overlap_detection_enabled = True
        
        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Create layout
        layout = QVBoxLayout(central_widget)
        
        # Create minimal toolbar
        toolbar_layout = QHBoxLayout()
        
        import_btn = QPushButton("Import Array")
        import_btn.clicked.connect(self.import_array_from_csv)
        toolbar_layout.addWidget(import_btn)
        
        import_bg_btn = QPushButton("Import Background")
        import_bg_btn.clicked.connect(self.import_background_image)
        toolbar_layout.addWidget(import_bg_btn)
        
        clear_bg_btn = QPushButton("Clear Background")
        clear_bg_btn.clicked.connect(self.clear_background_image)
        toolbar_layout.addWidget(clear_bg_btn)
        
        clear_btn = QPushButton("Clear")
        clear_btn.clicked.connect(self.clear_all)
        toolbar_layout.addWidget(clear_btn)
        
        grid_btn = QPushButton("Grid 250")
        grid_btn.clicked.connect(self.toggle_grid)
        toolbar_layout.addWidget(grid_btn)
        
        cut_btn = QPushButton("Cut")
        cut_btn.clicked.connect(self.perform_cut)
        toolbar_layout.addWidget(cut_btn)
        
        save_boxes_btn = QPushButton("Save Boxes")
        save_boxes_btn.clicked.connect(self.save_a1_box)
        toolbar_layout.addWidget(save_boxes_btn)
        
        report_btn = QPushButton("Report")
        report_btn.clicked.connect(self.create_shape_report)
        toolbar_layout.addWidget(report_btn)
        
        organize_btn = QPushButton("Organize")
        organize_btn.clicked.connect(self.toggle_shape_numbers)
        toolbar_layout.addWidget(organize_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        # Create graphics view
        self.cutter_view = CutterView(self)
        layout.addWidget(self.cutter_view)
        
        # Create menu bar
        self.create_menu_bar()
    
    def create_menu_bar(self):
        """Create minimal menu bar"""
        menu_bar = self.menuBar()
        
        # File menu
        file_menu = menu_bar.addMenu("File")
        
        import_action = QAction("Import Array", self)
        import_action.triggered.connect(self.import_array_from_csv)
        file_menu.addAction(import_action)
        
        load_bg_action = QAction("Load Background Image", self)
        load_bg_action.triggered.connect(self.load_background)
        file_menu.addAction(load_bg_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Organize menu
        organize_menu = menu_bar.addMenu("Organize")
        
        show_numbers_action = QAction("Show Shape Numbers", self)
        show_numbers_action.triggered.connect(self.toggle_shape_numbers)
        organize_menu.addAction(show_numbers_action)
    
    def load_background(self):
        """Load background image"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Load Background Image", "", 
            "Images (*.png *.jpg *.jpeg *.bmp *.gif);;All Files (*)"
        )
        if file_path:
            pixmap = QPixmap(file_path)
            if not pixmap.isNull():
                self.cutter_view.set_background_image(pixmap)
    
    def toggle_shape_numbers(self):
        """Toggle display of shape serial numbers on shapes"""
        self.cutter_view.toggle_shape_numbers()
    
    def clear_all(self):
        """Clear all shapes and cut lines"""
        self.cutter_view.clear_shapes()
        self.cutter_view.clear_cut_lines()
    
    def perform_cut(self):
        """Perform cut operation - fill all boxes that contain shapes with different colors"""
        self.cutter_view.fill_A1_and_A2_boxes()
        
        # Automatically show borders after cutting
        self.cutter_view.draw_red_green_border()
        
        # Fill all boxes with white after saving files
        self.cutter_view.fill_all_boxes_white()
        
        # Restore original fill colors for all shapes
        self.restore_original_colors()
    
    def export_svg(self):
        """Show borders of all colored blobs after cut operation"""
        self.cutter_view.draw_red_green_border()
    
    def restore_original_colors(self):
        """Restore original fill colors to all shapes and color their frames to match the fill color"""
        from PyQt5.QtGui import QColor
        
        print("Restoring original colors to shapes...")
        shapes_restored = 0
        
        # Get all items in the scene
        for item in self.cutter_view.scene.items():
            # Check if it's one of our shape types and has original color data
            if isinstance(item, (ScalableRectangle, ScalableTriangle)) and hasattr(item, 'original_fill_color'):
                original_fill_color = getattr(item, 'original_fill_color', '')
                original_frame_color = getattr(item, 'original_frame_color', '#8B4513')
                original_is_filled = getattr(item, 'original_is_filled', False)
                
                try:
                    # Handle fill color
                    frame_color_to_use = QColor(0, 0, 0)  # Default to black
                    
                    if original_is_filled and original_fill_color and original_fill_color.strip():
                        # Shape should be filled with original color
                        fill_color = QColor(original_fill_color)
                        if fill_color.isValid():
                            item.set_fill_color(fill_color)
                            frame_color_to_use = fill_color  # Frame color matches fill color
                            print(f"Restored fill color {original_fill_color} to shape {getattr(item, 'serial_number', 'unknown')}")
                        else:
                            # Invalid fill color, keep transparent
                            item.setBrush(QBrush(Qt.transparent))
                            print(f"Invalid fill color format: {original_fill_color}")
                    else:
                        # Shape should be transparent (not filled)
                        item.setBrush(QBrush(Qt.transparent))
                        # For transparent shapes, use the original frame color if available
                        if original_frame_color and original_frame_color.strip():
                            potential_frame_color = QColor(original_frame_color)
                            if potential_frame_color.isValid():
                                frame_color_to_use = potential_frame_color
                    
                    # Set frame color (either matching fill color or original frame color)
                    pen = QPen(frame_color_to_use, 0)  # Width 0 = cosmetic (always 1 pixel)
                    pen.setCosmetic(True)
                    item.setPen(pen)
                    
                    shapes_restored += 1
                    
                except Exception as e:
                    print(f"Error setting colors for shape {getattr(item, 'serial_number', 'unknown')}: {e}")
                    # Fallback to transparent with black frame
                    item.setBrush(QBrush(Qt.transparent))
                    item.setPen(QPen(QColor(0, 0, 0), 0))
        
        print(f"Restored original colors to {shapes_restored} shapes")

    def create_shape_report(self):
        """Create Excel reports - both general report and individual box reports"""
        try:
            from datetime import datetime
            import os
            import csv
            
            # Check if we have a currently imported CSV file
            if not self.current_csv_file:
                print("Error: No CSV file has been imported yet. Please import a CSV file first.")
                return
            
            # Try to import openpyxl for Excel file creation
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            except ImportError:
                print("Error: openpyxl library is required for Excel report generation.")
                print("Please install it using: pip install openpyxl")
                return
            
            # Dictionary to store shape counts (general report)
            shape_counts = {}
            
            # Dictionary to store shapes by box (box reports)
            box_shapes = {}  # box_name -> [(shape_type, color, count), ...]
            
            # Grid parameters for box calculation
            box_size = 250
            grid_cols = 6
            grid_rows = 6
            grid_offset_x = getattr(self.cutter_view, 'grid_offset_x', 0)
            grid_offset_y = getattr(self.cutter_view, 'grid_offset_y', 0)
            
            # Read data directly from the currently imported CSV file
            try:
                with open(self.current_csv_file, 'r', newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    
                    # Skip header row
                    header = next(reader, None)
                    if not header:
                        print("Error: Empty CSV file")
                        return
                    
                    # Process each row
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            if len(row) < 10:
                                print(f"Warning: Row {row_num} has insufficient data, skipping")
                                continue
                            
                            # Parse CSV data
                            shape_type = row[1]
                            x = float(row[2])
                            y = float(row[3])
                            width = float(row[4])
                            height = float(row[5])
                            fill_color = row[8] if row[8] else ""
                            is_filled = row[9].lower() in ('true', '1', 'yes') if row[9] else False
                            
                            # Determine shape type and size based on width and height
                            if shape_type == "Triangle":
                                # For triangles, use width as the size
                                size_key = f"{int(width)}X{int(width)} Triangle"
                            else:  # Rectangle
                                # For rectangles, distinguish between full and half rectangles
                                w = int(width)
                                h = int(height)
                                if w == h:
                                    # Square rectangle
                                    size_key = f"{w}X{h}"
                                else:
                                    # Half rectangle (width != height)
                                    size_key = f"{max(w, h)}X{min(w, h)}"
                            
                            # Determine color
                            if is_filled and fill_color and fill_color.strip():
                                color_hex = fill_color.upper()
                            else:
                                color_hex = "Transparent"
                            
                            # Create key for grouping (size_key, color)
                            key = (size_key, color_hex)
                            
                            # Count this shape for general report
                            if key in shape_counts:
                                shape_counts[key] += 1
                            else:
                                shape_counts[key] = 1
                            
                            # Determine which box this shape belongs to
                            shape_center_x = x + width / 2
                            shape_center_y = y + height / 2
                            
                            # Find the box that contains the shape's center
                            box_col = int((shape_center_x - grid_offset_x) / box_size)
                            box_row = int((shape_center_y - grid_offset_y) / box_size)
                            
                            # Debug: Print box calculation details for first few shapes
                            if row_num <= 5:
                                print(f"Debug Shape {row_num}: center=({shape_center_x:.1f}, {shape_center_y:.1f}), "
                                      f"grid_offset=({grid_offset_x}, {grid_offset_y}), "
                                      f"calculated_box=({box_col}, {box_row})")
                            
                            # Check if the shape is within the grid bounds
                            if 0 <= box_col < grid_cols and 0 <= box_row < grid_rows:
                                # Calculate box name (A1, B2, etc.)
                                col_letter = chr(ord('A') + box_col)
                                row_number = box_row + 1
                                box_name = f"{col_letter}{row_number}"
                                
                                # Add to box shapes dictionary
                                if box_name not in box_shapes:
                                    box_shapes[box_name] = {}
                                
                                if key in box_shapes[box_name]:
                                    box_shapes[box_name][key] += 1
                                else:
                                    box_shapes[box_name][key] = 1
                                    
                                # Debug: Print assignment for first few shapes
                                if row_num <= 5:
                                    print(f"Debug: Shape {row_num} assigned to box {box_name}")
                            else:
                                # Debug: Print shapes that fall outside grid
                                if row_num <= 10:
                                    print(f"Debug: Shape {row_num} outside grid bounds: box_col={box_col}, box_row={box_row}")
                                
                        except (ValueError, IndexError) as e:
                            print(f"Warning: Error parsing row {row_num}: {e}, skipping")
                            continue
                            
            except Exception as e:
                print(f"Error reading CSV file: {e}")
                return
            
            # Debug: Print summary of what was found
            print(f"Debug: Total unique shape types found: {len(shape_counts)}")
            print(f"Debug: Boxes with shapes: {list(box_shapes.keys())}")
            print(f"Debug: Total boxes with shapes: {len(box_shapes)}")
            for box_name, shapes in box_shapes.items():
                total_in_box = sum(shapes.values())
                print(f"Debug: Box {box_name} has {total_in_box} shapes ({len(shapes)} unique types)")
            
            # Show save directory dialog
            from PyQt5.QtWidgets import QFileDialog
            save_dir = QFileDialog.getExistingDirectory(
                self, "Select Directory for Reports", "",
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
            )
            
            if not save_dir:
                print("Report generation cancelled by user")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Create individual box reports first
            self.create_box_reports(box_shapes, save_dir, timestamp)
            
            # Create general report as sum of all box reports (calculate from box_shapes)
            general_shape_counts = {}
            for box_name, shapes in box_shapes.items():
                for key, count in shapes.items():
                    if key in general_shape_counts:
                        general_shape_counts[key] += count
                    else:
                        general_shape_counts[key] = count
            
            self.create_general_report(general_shape_counts, save_dir, timestamp)
            
            print(f"All reports created in: {save_dir}")
            print(f"General report + {len(box_shapes)} box reports generated")
            
        except Exception as e:
            print(f"Error creating shape reports: {e}")
    
    def create_general_report(self, shape_counts, save_dir, timestamp):
        """Create the general shape report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            import os
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "General Shape Report"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Write headers
            headers = ['Shape Type', 'Color', 'Count']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Sort by shape type, then color
            sorted_shapes = sorted(shape_counts.items(), key=lambda x: (x[0][0], x[0][1]))
            
            # Write data rows
            for row_idx, ((shape_type, color), count) in enumerate(sorted_shapes, 2):
                self.write_shape_row(ws, row_idx, shape_type, color, count, border)
            
            # Add total row
            self.add_total_row(ws, len(sorted_shapes), sum(shape_counts.values()), border)
            
            # Auto-adjust column widths
            self.adjust_column_widths(ws)
            
            # Save the workbook
            general_report_path = os.path.join(save_dir, f"general_shape_report_{timestamp}.xlsx")
            wb.save(general_report_path)
            
            print(f"General report created: {general_report_path}")
            
        except Exception as e:
            print(f"Error creating general report: {e}")
    
    def create_box_reports(self, box_shapes, save_dir, timestamp):
        """Create individual reports for each box"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            import os
            import traceback
            
            print(f"Debug: create_box_reports called with {len(box_shapes)} boxes")
            
            # Create a subdirectory for box reports
            box_reports_dir = os.path.join(save_dir, f"box_reports_{timestamp}")
            print(f"Debug: Creating box reports directory: {box_reports_dir}")
            
            if not os.path.exists(box_reports_dir):
                os.makedirs(box_reports_dir)
                print(f"Debug: Box reports directory created successfully")
            else:
                print(f"Debug: Box reports directory already exists")
            
            if not box_shapes:
                print("Debug: No box shapes data - box_shapes dictionary is empty")
                return
            
            for box_name, shapes in box_shapes.items():
                print(f"Debug: Creating report for box {box_name} with {len(shapes)} shape types")
                
                # Create workbook for this box
                wb = Workbook()
                ws = wb.active
                ws.title = f"Box {box_name} Report"
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                
                # Add box title
                title_cell = ws.cell(row=1, column=1, value=f"BOX {box_name} REPORT")
                title_cell.font = Font(bold=True, size=16, color="000000")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells('A1:C1')
                
                # Write headers
                headers = ['Shape Type', 'Color', 'Count']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Sort shapes by type, then color
                sorted_shapes = sorted(shapes.items(), key=lambda x: (x[0][0], x[0][1]))
                
                # Write data rows
                for row_idx, ((shape_type, color), count) in enumerate(sorted_shapes, 4):
                    self.write_shape_row(ws, row_idx, shape_type, color, count, border)
                
                # Add total row for this box
                total_shapes_in_box = sum(shapes.values())
                self.add_total_row(ws, len(sorted_shapes), total_shapes_in_box, border, start_row=4)
                
                # Auto-adjust column widths
                self.adjust_column_widths(ws)
                
                # Save the box report
                box_report_path = os.path.join(box_reports_dir, f"box_{box_name}_report_{timestamp}.xlsx")
                wb.save(box_report_path)
                
                print(f"Box {box_name} report created: {box_report_path}")
                
            print(f"Debug: Finished creating {len(box_shapes)} box reports in {box_reports_dir}")
                
        except Exception as e:
            print(f"Error creating box reports: {e}")
            import traceback
            traceback.print_exc()
    
    def write_shape_row(self, ws, row_idx, shape_type, color, count, border):
        """Write a single shape data row to the worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Shape Type column
        type_cell = ws.cell(row=row_idx, column=1, value=shape_type)
        type_cell.border = border
        type_cell.alignment = Alignment(horizontal="center")
        
        # Color column - apply the actual color as background
        color_cell = ws.cell(row=row_idx, column=2, value=color)
        color_cell.border = border
        color_cell.alignment = Alignment(horizontal="center")
        
        # Apply color formatting
        if color != "Transparent" and color.startswith("#") and len(color) == 7:
            try:
                # Remove # and use hex color for background
                hex_color = color[1:]  # Remove #
                color_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                color_cell.fill = color_fill
                
                # Calculate if we need light or dark text based on color brightness
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                brightness = (r * 0.299 + g * 0.587 + b * 0.114)
                
                if brightness < 128:  # Dark background, use white text
                    color_cell.font = Font(color="FFFFFF", bold=True)
                else:  # Light background, use black text
                    color_cell.font = Font(color="000000", bold=True)
                    
            except ValueError:
                # Invalid hex color, just use default formatting
                pass
        elif color == "Transparent":
            # For transparent, use a light gray background with "Transparent" text
            transparent_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            color_cell.fill = transparent_fill
            color_cell.font = Font(color="666666", italic=True)
        
        # Count column
        count_cell = ws.cell(row=row_idx, column=3, value=count)
        count_cell.border = border
        count_cell.alignment = Alignment(horizontal="center")
    
    def add_total_row(self, ws, num_data_rows, total_count, border, start_row=2):
        """Add a total row to the worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment
        
        total_row = start_row + num_data_rows + 1  # +1 for blank row
        
        # Create total row with merged cells for "TOTAL" label
        ws.merge_cells(f'A{total_row}:B{total_row}')
        total_label_cell = ws.cell(row=total_row, column=1, value="TOTAL")
        total_label_cell.font = Font(bold=True, size=12)
        total_label_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_label_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_label_cell.border = border
        
        # Apply border to merged cells
        for col in range(2, 3):  # column B
            cell = ws.cell(row=total_row, column=col)
            cell.border = border
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Add total count
        total_count_cell = ws.cell(row=total_row, column=3, value=total_count)
        total_count_cell.font = Font(bold=True, size=12)
        total_count_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_count_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_count_cell.border = border
    
    def adjust_column_widths(self, ws):
        """Auto-adjust column widths for the worksheet"""
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find a non-merged cell to get the column letter
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column = cell.column_letter
                        break
                except:
                    continue
            
            # If we couldn't find a column letter, skip this column
            if column is None:
                continue
                
            # Calculate max length for this column
            for cell in col:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            ws.column_dimensions[column].width = adjusted_width

    def save_a1_box(self):
        """Save PNG files for all boxes that contain shapes with 10-pixel margin"""
        if not self.cutter_view.grid_visible:
            print("Error: Grid must be visible to save boxes")
            return
        
        try:
            # Create blobs directory if it doesn't exist
            blobs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blobs")
            if not os.path.exists(blobs_dir):
                os.makedirs(blobs_dir)
            
            # Grid parameters
            box_size = 250
            margin = 20  # 20 pixel margin
            grid_cols = 6
            grid_rows = 6
            
            boxes_saved = 0
            
            # Check each box for shapes
            for row in range(grid_rows):
                for col in range(grid_cols):
                    # Calculate box position
                    box_x = self.cutter_view.grid_offset_x + (col * box_size)
                    box_y = self.cutter_view.grid_offset_y + (row * box_size)
                    box_rect = QRectF(box_x, box_y, box_size, box_size)
                    
                    # Check if this box contains any shapes
                    has_shapes = False
                    for item in self.cutter_view.scene.items():
                        if (item != self.cutter_view.background_item and 
                            item not in self.cutter_view.grid_items and 
                            item not in self.cutter_view.grid_labels and
                            item not in self.cutter_view.cut_lines and
                            item != self.cutter_view.grid_handle and
                            (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                            
                            # Check if shape overlaps with this box
                            shape_rect = item.sceneBoundingRect()
                            if box_rect.intersects(shape_rect):
                                has_shapes = True
                                break
                    
                    # If box contains shapes, save PNG
                    if has_shapes:
                        # Calculate box name (A1, B2, etc.)
                        col_letter = chr(ord('A') + col)
                        row_number = row + 1
                        box_name = f"{col_letter}{row_number}"
                        
                        # Define the capture area with margin
                        capture_x = box_x - margin
                        capture_y = box_y - margin
                        capture_width = box_size + (2 * margin)
                        capture_height = box_size + (2 * margin)
                        
                        # Temporarily hide unwanted items during PNG rendering
                        # Hide extra shape frames but keep blob borders, and hide circles/text
                        hidden_frames = []
                        for cut_item in self.cutter_view.cut_lines:
                            # Hide shape frame items (created by draw_shape_frames with z-value 1.5)
                            if (isinstance(cut_item, (QGraphicsRectItem, QGraphicsPolygonItem)) and
                                cut_item.pen().color() == QColor(0, 0, 0) and
                                cut_item.brush().color() == Qt.transparent and
                                hasattr(cut_item, 'zValue') and cut_item.zValue() == 1.5):
                                cut_item.setVisible(False)
                                hidden_frames.append(cut_item)
                            # Hide circles (QGraphicsEllipseItem from draw_red_green_border)
                            elif hasattr(cut_item, '__class__') and 'Ellipse' in cut_item.__class__.__name__:
                                cut_item.setVisible(False)
                                hidden_frames.append(cut_item)
                            # Hide text line items (created by draw_line_text with z-value 4)
                            elif (hasattr(cut_item, '__class__') and 'Line' in cut_item.__class__.__name__ and
                                  hasattr(cut_item, 'zValue') and cut_item.zValue() == 4):
                                cut_item.setVisible(False)
                                hidden_frames.append(cut_item)
                        
                        # Also hide any text items in the scene
                        from PyQt5.QtWidgets import QGraphicsTextItem, QGraphicsEllipseItem
                        hidden_text_and_circles = []
                        for item in self.cutter_view.scene.items():
                            if isinstance(item, (QGraphicsTextItem, QGraphicsEllipseItem)):
                                hidden_text_and_circles.append((item, item.isVisible()))
                                item.setVisible(False)
                        
                        # Also temporarily make shape frames transparent to match on-screen appearance
                        original_shape_pens = []
                        for item in self.cutter_view.scene.items():
                            if (isinstance(item, (ScalableRectangle, ScalableTriangle)) and
                                item != self.cutter_view.background_item and 
                                item not in self.cutter_view.grid_items and 
                                item not in self.cutter_view.grid_labels and
                                item not in self.cutter_view.cut_lines and
                                item != self.cutter_view.grid_handle):
                                # Store original pen and set transparent pen
                                original_shape_pens.append((item, item.pen()))
                                transparent_pen = QPen(Qt.transparent, 0)
                                transparent_pen.setCosmetic(True)
                                item.setPen(transparent_pen)
                        
                        # Create high-quality pixmap
                        pixmap = QPixmap(capture_width, capture_height)
                        pixmap.fill(Qt.white)  # White background
                        
                        # Create QPainter for high-quality rendering
                        from PyQt5.QtGui import QPainter
                        painter = QPainter(pixmap)
                        
                        # Enable high-quality rendering
                        painter.setRenderHint(QPainter.Antialiasing)
                        painter.setRenderHint(QPainter.TextAntialiasing)
                        painter.setRenderHint(QPainter.SmoothPixmapTransform)
                        
                        # Define the source rectangle (scene coordinates)
                        source_rect = QRectF(capture_x, capture_y, capture_width, capture_height)
                        
                        # Define the target rectangle (pixmap coordinates)
                        target_rect = QRectF(0, 0, capture_width, capture_height)
                        
                        # Render the scene area to the pixmap
                        self.cutter_view.scene.render(painter, target_rect, source_rect)
                        painter.end()
                        
                        # Restore visibility of all hidden items
                        for frame_item in hidden_frames:
                            frame_item.setVisible(True)
                        
                        # Restore visibility of text and circles
                        for item, was_visible in hidden_text_and_circles:
                            item.setVisible(was_visible)
                        
                        # Restore original shape pens
                        for item, original_pen in original_shape_pens:
                            item.setPen(original_pen)
                        
                        # Save PNG file to blobs directory
                        png_filename = f"{box_name}_box.png"
                        png_path = os.path.join(blobs_dir, png_filename)
                        
                        success = pixmap.save(png_path, "PNG")
                        
                        if success:
                            print(f"Box {box_name} saved as PNG: {png_path}")
                            boxes_saved += 1
                        else:
                            print(f"Error: Failed to save box {box_name} to {png_path}")
                        
                        # Save CSV file with shapes in this box using box-relative coordinates
                        self.save_box_shapes_csv(box_name, box_x, box_y, box_size, blobs_dir)
            
            print(f"Successfully saved {boxes_saved} box PNG files to blobs directory")
            
        except Exception as e:
            print(f"Error saving box PNG files: {e}")
    
    def save_box_shapes_csv(self, box_name, box_x, box_y, box_size, blobs_dir):
        """Save shapes in a specific box to a CSV file with box-relative coordinates"""
        try:
            import csv
            
            # Calculate box index from box name (A1=0, A2=1, ..., B1=6, B2=7, etc.)
            col_letter = box_name[0]
            row_number = int(box_name[1:])
            col = ord(col_letter) - ord('A')
            row = row_number - 1
            box_index = row * 6 + col
            
            # Use stored inclusion data from the first shape selection
            box_shapes = []
            if box_index in self.cutter_view.box_inclusion_data:
                box_shapes = self.cutter_view.box_inclusion_data[box_index]
                print(f"Using stored inclusion data for box {box_name} (index {box_index}): {len(box_shapes)} shapes")
            else:
                print(f"No inclusion data found for box {box_name} (index {box_index}) - using fallback calculation")
                # Fallback to original calculation if no stored data
                box_rect = QRectF(box_x, box_y, box_size, box_size)
                for item in self.cutter_view.scene.items():
                    if (item != self.cutter_view.background_item and 
                        item not in self.cutter_view.grid_items and 
                        item not in self.cutter_view.grid_labels and
                        item not in self.cutter_view.cut_lines and
                        item != self.cutter_view.grid_handle and
                        (isinstance(item, ScalableRectangle) or isinstance(item, ScalableTriangle))):
                        
                        shape_rect = item.sceneBoundingRect()
                        
                        # Calculate the intersection area between shape and box
                        intersection = box_rect.intersected(shape_rect)
                        if not intersection.isEmpty():
                            # Calculate overlap percentage relative to the shape size
                            shape_area = shape_rect.width() * shape_rect.height()
                            intersection_area = intersection.width() * intersection.height()
                            overlap_percentage = (intersection_area / shape_area) * 100 if shape_area > 0 else 0
                            
                            # Use 25% overlap threshold (same as inclusion logic)
                            if overlap_percentage >= 25.0:
                                box_shapes.append(item)
            
            if not box_shapes:
                print(f"No shapes found in box {box_name} - skipping CSV creation")
                return
            
            # Sort shapes from top to bottom (by Y coordinate)
            box_shapes.sort(key=lambda item: item.pos().y())
            
            # Create CSV file for this box
            csv_filename = f"{box_name}_shapes.csv"
            csv_path = os.path.join(blobs_dir, csv_filename)
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow([
                    'Serial_Number', 'Shape_Type', 'X', 'Y', 'Width', 'Height', 
                    'Rotation', 'Frame_Color', 'Fill_Color', 'Is_Filled'
                ])
                
                # Write shape data with box-relative coordinates (top-left corner as 0,0)
                for item in box_shapes:
                    # Get shape properties
                    serial_number = getattr(item, 'serial_number', 0)
                    
                    # Determine shape type and get proper dimensions
                    if isinstance(item, ScalableTriangle):
                        shape_type = "Triangle"
                        # For triangles, use the stored size parameter for both width and height
                        size = getattr(item, 'size', 0)
                        width = size
                        height = size
                    else:
                        shape_type = "Rectangle"
                        # For rectangles, get dimensions from the internal rect
                        rect = item.rect()
                        width = rect.width()
                        height = rect.height()
                    
                    # Get position relative to origin (175 pixels left and 135 pixels above box top-left)
                    shape_pos = item.pos()
                    relative_x = shape_pos.x() - (box_x - 175)
                    relative_y = shape_pos.y() - (box_y - 135)
                    
                    # Get rotation
                    rotation = getattr(item, 'current_rotation', 0)
                    
                    # Get original colors
                    original_fill_color = getattr(item, 'original_fill_color', '')
                    original_frame_color = getattr(item, 'original_frame_color', '#8B4513')
                    original_is_filled = getattr(item, 'original_is_filled', False)
                    
                    # Write row
                    writer.writerow([
                        serial_number,
                        shape_type,
                        f"{relative_x:.2f}",
                        f"{relative_y:.2f}",
                        f"{width:.2f}",
                        f"{height:.2f}",
                        f"{rotation:.2f}",
                        original_frame_color,
                        original_fill_color,
                        original_is_filled
                    ])
            
            print(f"Box {box_name} shapes saved to CSV: {csv_path} ({len(box_shapes)} shapes)")
            
        except Exception as e:
            print(f"Error saving CSV for box {box_name}: {e}")
    
    def toggle_grid(self):
        """Toggle the 250x250 grid on/off"""
        if self.cutter_view.grid_visible:
            self.cutter_view.clear_grid()
        else:
            self.cutter_view.create_grid()
    
    def import_array_from_csv(self):
        """Import shape data from CSV file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Array from CSV", "", 
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            # Store the current CSV file path for reports
            self.current_csv_file = file_path
            
            try:
                shapes_created = 0
                self.cutter_view.clear_shapes()
                
                with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    
                    # Skip header row
                    header = next(reader, None)
                    if not header:
                        print("Error: Empty CSV file")
                        return
                    
                    # Process each row
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            if len(row) < 10:
                                print(f"Warning: Row {row_num} has insufficient data, skipping")
                                continue
                            
                            # Parse CSV data
                            serial_number = int(row[0]) if row[0] else 0
                            shape_type = row[1]
                            x = float(row[2])
                            y = float(row[3])
                            width = float(row[4])
                            height = float(row[5])
                            rotation = float(row[6]) if row[6] else 0
                            frame_color = row[7] if row[7] else "#8B4513"
                            fill_color = row[8] if row[8] else ""
                            is_filled = row[9].lower() in ('true', '1', 'yes') if row[9] else False
                            
                            # Create shape
                            if shape_type == "Triangle":
                                shape = ScalableTriangle(x, y, width)
                            else:
                                shape = ScalableRectangle(x, y, width, height)
                            
                            shape.serial_number = serial_number
                            
                            # Store original colors for later restoration
                            shape.original_fill_color = fill_color if fill_color else ""
                            shape.original_frame_color = frame_color if frame_color else "#8B4513"
                            shape.original_is_filled = is_filled
                            
                            # Set rotation if specified
                            if rotation != 0:
                                shape.current_rotation = rotation
                                shape.setRotation(rotation)
                            
                            # Always keep shapes transparent with black frame - ignore saved colors
                            # This ensures all shapes are displayed as transparent regardless of CSV data
                            
                            self.cutter_view.add_shape(shape)
                            shapes_created += 1
                            
                        except (ValueError, IndexError) as e:
                            print(f"Warning: Error parsing row {row_num}: {e}, skipping")
                            continue
                
                print(f"Successfully imported {shapes_created} shapes from: {file_path}")
                
                # Center view on imported shapes if any were created
                if shapes_created > 0:
                    self.center_on_content()
                
            except Exception as e:
                print(f"Error importing CSV file: {e}")
    
    def center_on_content(self):
        """Center the view on all shapes"""
        # Get bounding rectangle of all items (excluding background)
        items_rect = None
        for item in self.cutter_view.scene.items():
            if item != self.cutter_view.background_item:
                if items_rect is None:
                    items_rect = item.sceneBoundingRect()
                else:
                    items_rect = items_rect.united(item.sceneBoundingRect())
        
        if items_rect is not None:
            # Add some padding
            padding = 50
            items_rect.adjust(-padding, -padding, padding, padding)
            self.cutter_view.fitInView(items_rect, Qt.KeepAspectRatio)
            # Don't zoom too much - limit the scale
            current_scale = self.cutter_view.transform().m11()
            if current_scale > 2.0:  # If zoomed in too much, zoom out a bit
                self.cutter_view.scale(0.5, 0.5)
    
    def import_background_image(self):
        """Import a background image"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Background Image", "", 
            "Image Files (*.png *.jpg *.jpeg *.bmp *.gif *.tiff);;All Files (*)"
        )
        if file_path:
            try:
                # Load the image
                pixmap = QPixmap(file_path)
                if pixmap.isNull():
                    print(f"Error: Could not load image from {file_path}")
                    return
                
                # Set the background image
                self.cutter_view.set_background_image(pixmap)
                
                # Position the background image at (155, 115)
                if self.cutter_view.background_item:
                    self.cutter_view.background_item.setPos(155, 115)
                
                print(f"Successfully imported background image: {file_path}")
                print(f"Background positioned at (155, 115)")
                
                # Optionally center the view on the background
                self.cutter_view.centerOn(self.cutter_view.background_item)
                
            except Exception as e:
                print(f"Error importing background image: {e}")
    
    def clear_background_image(self):
        """Clear the background image"""
        if self.cutter_view.background_item:
            self.cutter_view.scene.removeItem(self.cutter_view.background_item)
            self.cutter_view.background_item = None
            print("Background image cleared")
        else:
            print("No background image to clear")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CutterWindow()
    window.show()
    sys.exit(app.exec_())
